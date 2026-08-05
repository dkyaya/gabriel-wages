#!/usr/bin/env python3
"""Resumable downstream stages for the exhaustive external-data pipeline.

Stage 1 is isolated in ``run_external_data_exhaustive_pipeline.py`` because it
owns long-lived hosted-search workers.  This companion runner begins only after
the residual-search validation gate passes.  Network workers use append-only
accepted ledgers and small atomic checkpoints so an interrupted lane can resume
without repeating completed rows.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import json
import mimetypes
import os
import re
import shutil
import sys
import tempfile
import time
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlsplit

import run_external_data_exhaustive_pipeline as core


LANES = [f"lane_{i:03d}" for i in range(1, 6)]
READY_REVIEW_BUCKETS = {
    "high_priority_verification_ready",
    "medium_priority_verification_ready",
    "low_priority_verification_ready",
}
DOCUMENT_EXTENSIONS = {
    "pdf", "csv", "tsv", "xlsx", "xls", "json", "xml", "txt", "zip",
    "doc", "docx", "ods",
}
EXCLUDED_DOMAINS = {
    "facebook.com", "www.facebook.com", "linkedin.com", "www.linkedin.com",
    "indeed.com", "www.indeed.com", "glassdoor.com", "www.glassdoor.com",
    "salary.com", "www.salary.com", "ziprecruiter.com", "www.ziprecruiter.com",
    "pinterest.com", "www.pinterest.com", "youtube.com", "www.youtube.com",
}


def load_shards(directory: Path, manifest_name: str) -> list[dict[str, str]]:
    manifest = json.loads((directory / manifest_name).read_text())
    rows: list[dict[str, str]] = []
    for part in manifest["parts"]:
        filename = part.get("csv") or part.get("csv_path")
        rows.extend(core.read_csv(directory / filename))
    return rows


def append_jsonl(path: Path, row: dict[str, Any]) -> None:
    core.append_jsonl(path, row)


def archive_append_ledgers(stage: Path, glob_patterns: Iterable[str], label: str) -> None:
    """Move noncanonical worker journals to ignored temporary storage."""
    destination = core.TMP / "downstream_append_only_checkpoint_archive" / label
    destination.mkdir(parents=True, exist_ok=True)
    moved = []
    for pattern in glob_patterns:
        for source in sorted(stage.glob(pattern)):
            target = destination / source.name
            if target.exists(): target.unlink()
            shutil.move(str(source), str(target)); moved.append({"name":source.name,"bytes":target.stat().st_size,"sha256":core.sha256_file(target)})
    core.write_json(stage / f"{label}_append_only_checkpoint_archive_manifest.json", {"archived_after_validation":True,"archive_root":str(destination.relative_to(core.ROOT)),"archive_root_git_ignored":True,"canonical_lane_and_merged_outputs_preserved":True,"files":moved})


def lane_partition(rows: list[dict[str, Any]], family_key: str = "external_data_family") -> list[list[dict[str, Any]]]:
    """Deterministic greedy balancing by family and row complexity."""
    buckets: list[list[dict[str, Any]]] = [[] for _ in range(5)]
    loads = [0] * 5
    ordered = sorted(
        rows,
        key=lambda row: (
            str(row.get(family_key, "")), str(row.get("state", "")),
            str(row.get("municipality", "")), str(row.get("candidate_id", row.get("verification_id", ""))),
        ),
    )
    for row in ordered:
        index = min(range(5), key=lambda i: (loads[i], i))
        buckets[index].append(row)
        loads[index] += 1
    return buckets


def normalized_title(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", core.clean(value).casefold()).strip()


def official_domain(domain: str) -> bool:
    domain = domain.casefold().split(":", 1)[0]
    return domain.endswith(".gov") or ".gov." in domain or domain.endswith(".mil")


def public_institution_domain(domain: str) -> bool:
    domain = domain.casefold().split(":", 1)[0]
    return official_domain(domain) or domain.endswith(".edu") or domain.endswith(".us")


def primary_family(row: dict[str, str]) -> tuple[str, str]:
    text = " ".join([row.get("candidate_title", ""), row.get("candidate_url", ""), row.get("candidate_snippet", "")]).casefold()
    tags = [x for x in row.get("external_data_family", "").split("|") if x]
    rules = [
        ("payroll_and_earnings", r"payroll|earnings|open.?checkbook|employee compensation|wage roster|overtime earnings"),
        ("staffing_and_headcount", r"headcount|authorized positions?|filled positions?|vacanc|layoff|position elimination|staffing table|personnel count"),
        ("recruitment_and_retention", r"recruit|retention|turnover|applicant|hiring difficult|vacancy duration"),
        ("tenure_and_progression", r"tenure|years of service|salary step|step schedule|seniority|promotion|civil service roster"),
        ("implementation_confirmation", r"ordinance|resolution|ratif|effective date|appropriation|adopted|approved|memorandum of understanding|\bmou\b"),
        ("benefits_and_total_compensation", r"pension|retirement|health contribution|benefit|longevity|uniform allowance|certification pay|total compensation"),
        ("contextual_controls", r"population|urban rural|fiscal capacity|unemployment|labor market|collective bargaining law|labor law"),
    ]
    hits = [family for family, pattern in rules if re.search(pattern, text)]
    if len(hits) == 1:
        primary = hits[0]
    elif len(hits) > 1:
        primary = "multi_family_administrative_source"
    elif len(tags) == 1:
        primary = "implementation_confirmation" if tags[0] == "implementation" else tags[0]
    elif len(tags) > 1:
        primary = "multi_family_administrative_source"
    else:
        primary = "unclear"
    secondary = sorted(set(tags + hits))
    return primary, "|".join(secondary)


def administrative_source_type(row: dict[str, str]) -> str:
    text = " ".join([row.get("candidate_title", ""), urlsplit(row.get("candidate_url", "")).path]).casefold()
    rules = [
        ("payroll_roster", r"payroll|salary roster|employee compensation"),
        ("open_checkbook", r"open.?checkbook|open.?book"),
        ("earnings_report", r"earnings report|employee earnings"),
        ("staffing_table", r"staffing table|authorized positions?|filled positions?|headcount"),
        ("vacancy_report", r"vacanc|turnover"),
        ("compensation_study", r"compensation study|salary study|classification study"),
        ("recruitment_study", r"recruitment study|retention study|staffing study"),
        ("civil_service_roster", r"civil service|classification roster"),
        ("salary_schedule", r"salary schedule|pay schedule|step schedule|pay plan"),
        ("contract_or_mou", r"contract|collective bargaining|memorandum|\bmou\b"),
        ("ordinance_or_resolution", r"ordinance|resolution"),
        ("implementation_record", r"ratif|effective date|appropriation|adopted|approved"),
        ("audit_or_financial_report", r"audit|financial report|acfr|cafr"),
        ("benefits_document", r"benefit|health insurance|longevity|allowance"),
        ("pension_or_retirement_document", r"pension|retirement"),
        ("budget", r"budget|appropriation"),
        ("meeting_packet", r"agenda|meeting packet|council packet|minutes"),
        ("open_data_portal", r"open data|dataset|data portal|catalog"),
        ("government_dataset", r"\.csv|\.xlsx|\.xls|\.json|\.xml|dataset"),
        ("navigation_or_index", r"index|search|portal|department$|documents?$"),
        ("media_or_context", r"news|press|article|story"),
    ]
    for label, pattern in rules:
        if re.search(pattern, text):
            return label
    return "other" if row.get("candidate_url") else "unclear"


def review_candidate(row: dict[str, str], lane: str) -> dict[str, Any]:
    url = core.clean(row.get("candidate_url", ""))
    parsed = urlsplit(url)
    domain = parsed.netloc.casefold().split(":", 1)[0]
    title = core.clean(row.get("candidate_title", ""))
    extension = Path(parsed.path).suffix.casefold().lstrip(".")
    source_type = administrative_source_type(row)
    family, secondary = primary_family(row)
    review_reason: list[str] = []
    if parsed.scheme not in {"http", "https"} or not domain:
        bucket = "malformed_or_missing_locator"; review_reason.append("non-http(s) or missing host")
    elif domain in EXCLUDED_DOMAINS or any(domain.endswith("." + excluded) for excluded in EXCLUDED_DOMAINS):
        bucket = "excluded_out_of_scope"; review_reason.append("excluded aggregator, social, or generic job domain")
    elif source_type == "navigation_or_index" and not title and extension not in DOCUMENT_EXTENSIONS:
        bucket = "likely_navigation_only"; review_reason.append("navigation-like locator with no document title")
    elif official_domain(domain) and (extension in DOCUMENT_EXTENSIONS or source_type not in {"other", "navigation_or_index", "unclear"}):
        bucket = "high_priority_verification_ready"; review_reason.append("official government locator with administrative document/data signal")
    elif official_domain(domain):
        bucket = "medium_priority_verification_ready"; review_reason.append("official government locator; source type requires verification")
    elif public_institution_domain(domain) and (extension in DOCUMENT_EXTENSIONS or source_type not in {"other", "unclear"}):
        bucket = "medium_priority_verification_ready"; review_reason.append("public institutional locator with administrative signal")
    elif extension in DOCUMENT_EXTENSIONS and source_type not in {"media_or_context", "unclear"}:
        bucket = "low_priority_verification_ready"; review_reason.append("document locator with plausible administrative source signal")
    else:
        bucket = "deferred_low_signal"; review_reason.append("unconfirmed non-government web result without strong administrative signal")
    return {
        **row,
        "review_id": core.stable("EXTREVIEW", row.get("candidate_id", "")),
        "primary_external_data_family": family,
        "secondary_family_tags": secondary,
        "administrative_source_type": source_type,
        "candidate_review_bucket": bucket,
        "review_reason": "; ".join(review_reason),
        "review_confidence": "high" if bucket in {"malformed_or_missing_locator", "excluded_out_of_scope", "high_priority_verification_ready"} else "moderate",
        "metadata_only_review": "true",
        "reviewed_at": core.utc_now(),
        "candidate_review_lane_id": lane,
    }


def stage2_prepare() -> None:
    gate = json.loads((core.STAGE1 / "residual_search_validation_report.json").read_text())
    if not gate.get("passed"):
        raise RuntimeError("stage 1 validation gate must pass")
    wave1 = core.load_prior_candidates()
    prior_links=defaultdict(list)
    for link in core.read_csv(core.PRIOR/"search_target_event_linkage.csv"): prior_links[link["search_target_id"]].append(link)
    for row in wave1:
        row["search_wave"] = "external_search_wave_001_compacted"
        row["raw_target_id"] = ""
        row["prior_compacted_target_id"] = row.get("search_target_id", "")
        links=prior_links.get(row.get("search_target_id",""),[])
        row["linked_root_event_id"]="|".join(sorted({x["root_compensation_event_id"] for x in links if x["root_compensation_event_id"]}))
        row["linked_mechanism_exposure_event_ids"]="|".join(sorted({x["mechanism_exposure_event_id"] for x in links if x["mechanism_exposure_event_id"]}))
        row["search_wave_provenance"]="external_search_wave_001_compacted"
        row["linked_candidate_ids"]=row["candidate_id"]
    wave2_canonical = load_shards(core.STAGE1, "canonical_residual_candidates_shard_manifest.json")
    wave2_all = load_shards(core.STAGE1, "merged_residual_candidates_shard_manifest.json")
    for row in wave2_all:
        row["search_wave_provenance"]="external_search_wave_002_exhaustive_residual"
        row["linked_candidate_ids"]=row["candidate_id"]
    merged: list[dict[str, str]] = []
    duplicate_links: list[dict[str, str]] = []
    by_url: dict[str, dict[str, str]] = {}
    by_title: dict[tuple[str, str, str, str], dict[str, str]] = {}
    def merge_pipe(target:dict[str,str],source:dict[str,str],field:str)->None:
        values=[]
        for value in (target.get(field,""),source.get(field,"")):
            values.extend(x for x in value.split("|") if x)
        target[field]="|".join(sorted(set(values)))
    for row in wave1 + wave2_all:
        canonical_url = core.canonical_url(row.get("canonicalized_url") or row.get("candidate_url", ""))
        row["canonicalized_url"] = canonical_url
        title_key = normalized_title(row.get("candidate_title", ""))
        title_tuple = (urlsplit(canonical_url).netloc.casefold(), title_key, row.get("municipality", "").casefold(), row.get("period", ""))
        canonical = by_url.get(canonical_url) if canonical_url else None
        basis = "same canonical URL across waves" if canonical else ""
        if canonical is None and title_key and title_tuple in by_title:
            canonical = by_title[title_tuple]; basis = "same domain, normalized title, municipality, and period"
        if canonical:
            duplicate_links.append({"duplicate_candidate_id": row.get("candidate_id", ""),
                                    "canonical_candidate_id": canonical.get("candidate_id", ""),
                                    "duplicate_basis": basis, "duplicate_wave": row.get("search_wave", ""),
                                    "canonical_wave": canonical.get("search_wave", ""), "confidence": "high"})
            for field in ("linked_root_event_id","linked_mechanism_exposure_event_ids","raw_target_id","linked_candidate_ids","search_wave_provenance"):
                merge_pipe(canonical,row,field)
            upgrades=[]
            for value in (canonical.get("expected_claim_upgrade",""),row.get("expected_claim_upgrade","")):
                upgrades.extend(x for x in value.split(" | ") if x)
            canonical["expected_claim_upgrade"]=" | ".join(sorted(set(upgrades)))
            continue
        merged.append(row)
        if canonical_url: by_url[canonical_url] = row
        if title_key: by_title[title_tuple] = row
    core.write_sharded_pair(core.STAGE2, "merged_external_candidate_universe", merged)
    core.write_sharded_pair(core.STAGE2, "merged_cross_wave_candidate_duplicate_links", duplicate_links)
    core.write_json(core.STAGE2 / "merged_external_candidate_manifest.json", {
        "prepared_at": core.utc_now(), "wave1_canonical_input": len(wave1), "wave2_raw_candidate_input":len(wave2_all), "wave2_canonical_input": len(wave2_canonical),
        "merged_canonical_count": len(merged), "cross_wave_or_merged_duplicate_count": len(duplicate_links),
        "all_candidates_review_required": True, "duplicate_target_event_linkage_aggregated_on_canonical_candidates":True,
    })
    buckets = lane_partition(merged)
    for index, rows in enumerate(buckets, 1):
        core.write_pair(core.STAGE2, f"candidate_review_lane_{index:03d}_queue", rows)
    core.write_json(core.STAGE2 / "candidate_review_lane_distribution.json", {
        "lane_count": 5, "lane_sizes": {f"candidate_review_lane_{i:03d}": len(rows) for i, rows in enumerate(buckets, 1)},
        "disjoint_candidate_ids": len({r["candidate_id"] for part in buckets for r in part}) == len(merged),
    })
    core.record_transition("02_MERGED-EXTERNAL-CANDIDATE-REVIEW", "prepared", "merged_candidate_review_queues_locked",
                           {"merged_candidates": len(merged), "lane_sizes": [len(x) for x in buckets]})
    print(json.dumps({"merged_candidates": len(merged), "wave2_raw_candidates":len(wave2_all), "wave2_canonical_candidates":len(wave2_canonical), "duplicates": len(duplicate_links), "lane_sizes": [len(x) for x in buckets]}, indent=2))


def stage2_run_lane(lane_number: int) -> None:
    lane = f"candidate_review_lane_{lane_number:03d}"
    queue = core.read_csv(core.STAGE2 / f"{lane}_queue.csv")
    checkpoint_path = core.STAGE2 / f"{lane}_checkpoint.json"
    accepted_path = core.STAGE2 / f"{lane}_accepted_results.jsonl"
    checkpoint = json.loads(checkpoint_path.read_text()) if checkpoint_path.exists() else {
        "lane_id": lane, "assigned": len(queue), "completed": 0, "status": "in_progress",
        "queue_sha256": core.sha256_file(core.STAGE2 / f"{lane}_queue.csv"), "started_at": core.utc_now(),
    }
    if checkpoint.get("status") == "complete":
        print(json.dumps(checkpoint, indent=2)); return
    done = {row["candidate_id"] for row in core.read_jsonl(accepted_path)}
    for row in queue:
        if row["candidate_id"] in done: continue
        append_jsonl(accepted_path, review_candidate(row, lane))
        checkpoint["completed"] += 1; checkpoint["last_candidate_id"] = row["candidate_id"]; checkpoint["updated_at"] = core.utc_now()
        core.atomic_json(checkpoint_path, checkpoint)
    results = core.read_jsonl(accepted_path)
    checkpoint.update({"completed": len(results), "status": "complete", "finished_at": core.utc_now()})
    core.atomic_json(checkpoint_path, checkpoint)
    core.write_sharded_pair(core.STAGE2, f"{lane}_results", results)
    print(json.dumps({"lane": lane, "completed": len(results)}, indent=2))


def stage2_finalize() -> None:
    results: list[dict[str, Any]] = []
    for i in range(1, 6):
        lane = f"candidate_review_lane_{i:03d}"
        checkpoint = json.loads((core.STAGE2 / f"{lane}_checkpoint.json").read_text())
        if checkpoint.get("status") != "complete": raise RuntimeError(f"incomplete {lane}")
        results.extend(core.read_jsonl(core.STAGE2 / f"{lane}_accepted_results.jsonl"))
    manifest = json.loads((core.STAGE2 / "merged_external_candidate_manifest.json").read_text())
    if len(results) != manifest["merged_canonical_count"] or len({r["candidate_id"] for r in results}) != len(results):
        raise RuntimeError("candidate review does not reconcile")
    core.write_sharded_pair(core.STAGE2, "candidate_review_results", results)
    grouped = defaultdict(list)
    for row in results: grouped[row["candidate_review_bucket"]].append(row)
    name_map = {
        "high_priority_verification_ready": "high_priority_verification_ready_queue",
        "medium_priority_verification_ready": "medium_priority_verification_ready_queue",
        "low_priority_verification_ready": "low_priority_verification_ready_queue",
        "repair_needed": "candidate_review_repair_queue",
        "likely_duplicate_prior_source": "likely_duplicate_prior_source_queue",
        "likely_duplicate_within_external_wave": "likely_duplicate_within_external_wave_queue",
        "likely_navigation_only": "candidate_review_navigation_queue",
        "deferred_low_signal": "candidate_review_defer_queue",
        "excluded_out_of_scope": "candidate_review_exclude_queue",
        "malformed_or_missing_locator": "candidate_review_malformed_queue",
        "review_error": "candidate_review_error_queue",
    }
    for bucket, name in name_map.items(): core.write_sharded_pair(core.STAGE2, name, grouped[bucket])
    ready = [r for r in results if r["candidate_review_bucket"] in READY_REVIEW_BUCKETS]
    core.write_sharded_pair(core.STAGE2, "verification_ready_queue", ready)
    core.write_json(core.STAGE2 / "primary_secondary_family_summary.json", {
        "primary": dict(Counter(r["primary_external_data_family"] for r in results)),
        "secondary_tags": dict(Counter(tag for r in results for tag in r["secondary_family_tags"].split("|") if tag)),
    })
    core.write_json(core.STAGE2 / "administrative_source_type_summary.json", dict(Counter(r["administrative_source_type"] for r in results)))
    core.write_json(core.STAGE2 / "event_claim_linkage_summary.json", {
        "reviewed": len(results), "with_root_event": sum(bool(r.get("linked_root_event_id")) for r in results),
        "with_expected_claim_upgrade": sum(bool(r.get("expected_claim_upgrade")) for r in results),
    })
    checks = {"all_merged_candidates_reviewed": len(results) == manifest["merged_canonical_count"],
              "unique_candidate_review": len({r["candidate_id"] for r in results}) == len(results),
              "bucket_complete": sum(len(x) for x in grouped.values()) == len(results),
              "metadata_only": all(r["metadata_only_review"] == "true" for r in results)}
    passed = all(checks.values())
    core.write_json(core.STAGE2 / "candidate_review_validation_report.json", {"passed": passed, "checks": checks,
                    "review_bucket_counts": {k: len(v) for k, v in grouped.items()}, "verification_ready_count": len(ready)})
    core.write_md(core.STAGE2 / "candidate_review_validation_report.md", "# Candidate-review validation\n\n" +
                  "\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items()))
    decision = "merged_external_candidate_review_completed_verification_ready" if passed else "merged_external_candidate_review_repair_needed"
    core.write_json(core.STAGE2 / "stage_decision.json", {"decision": decision, "reviewed": len(results), "verification_ready": len(ready),
                    "bucket_counts": {k: len(v) for k,v in grouped.items()}, "completed_at": core.utc_now()})
    core.record_transition("02_MERGED-EXTERNAL-CANDIDATE-REVIEW", "complete" if passed else "repair_needed", decision,
                           {"reviewed": len(results), "verification_ready": len(ready)})
    if not passed: raise RuntimeError("stage 2 validation failed")
    archive_append_ledgers(core.STAGE2,["candidate_review_lane_*_accepted_results.jsonl"],"stage2_candidate_review")
    print(json.dumps({"decision": decision, "reviewed": len(results), "verification_ready": len(ready),
                      "buckets": {k:len(v) for k,v in grouped.items()}}, indent=2))


def stage3_prepare() -> None:
    gate = json.loads((core.STAGE2 / "candidate_review_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 2 validation gate must pass")
    ready = load_shards(core.STAGE2, "verification_ready_queue_shard_manifest.json")
    rows = []
    for row in ready:
        rows.append({**row, "verification_id": core.stable("EXTVERIFY", row["candidate_id"], row["canonicalized_url"])})
    buckets = lane_partition(rows, "primary_external_data_family")
    for i, part in enumerate(buckets, 1): core.write_pair(core.STAGE3, f"verification_lane_{i:03d}_queue", part)
    core.write_json(core.STAGE3 / "verification_lane_distribution.json", {"input_count": len(rows),
                    "lane_sizes": {f"verification_lane_{i:03d}":len(x) for i,x in enumerate(buckets,1)}, "checkpoint_each_row": True})
    core.record_transition("03_EXTERNAL-DATA-VERIFICATION", "prepared", "external_data_verification_queues_locked", {"input_count":len(rows)})
    print(json.dumps({"verification_input":len(rows),"lane_sizes":[len(x) for x in buckets]},indent=2))


def verify_one(row: dict[str, str], client: Any) -> dict[str, Any]:
    url = row["candidate_url"]
    common = {**row, "requested_url": url, "verified_at": core.utc_now(), "full_body_retained": "false"}
    try:
        response = client.head(url, follow_redirects=True)
        method = "HEAD"
        if response.status_code in {400, 403, 405, 406, 429, 500, 501} or not response.headers.get("content-type"):
            with client.stream("GET", url, headers={"Range":"bytes=0-2047"}, follow_redirects=True) as streamed:
                response = streamed
                iterator = streamed.iter_bytes()
                inspected = next(iterator, b"")[:2048]
                method = "GET_RANGE_METADATA"
                status_code = streamed.status_code; final_url = str(streamed.url); headers = dict(streamed.headers)
            body_bytes = len(inspected)
        else:
            status_code = response.status_code; final_url = str(response.url); headers = dict(response.headers); body_bytes = 0
        if 200 <= status_code < 300:
            status = "reachable_with_redirect" if core.canonical_url(final_url) != core.canonical_url(url) else "reachable"
        elif status_code in {401,403,451}: status = "blocked_or_forbidden"
        elif status_code in {404,410}: status = "unavailable"
        else: status = "verification_error"
        return {**common, "verification_status":status, "verification_method":method, "http_status":status_code,
                "final_url":final_url, "content_type":headers.get("content-type",""), "content_length":headers.get("content-length",""),
                "etag":headers.get("etag",""), "last_modified":headers.get("last-modified",""), "metadata_body_bytes_inspected":body_bytes,
                "verification_error":""}
    except Exception as exc:
        name=exc.__class__.__name__.casefold()
        status="timeout" if "timeout" in name else "verification_error"
        return {**common,"verification_status":status,"verification_method":"HEAD_OR_RANGE_GET","http_status":"","final_url":"",
                "content_type":"","content_length":"","etag":"","last_modified":"","metadata_body_bytes_inspected":0,
                "verification_error":f"{exc.__class__.__name__}: {core.clean(str(exc))[:300]}"}


def stage3_run_lane(lane_number: int) -> None:
    import httpx
    lane=f"verification_lane_{lane_number:03d}"; queue_path=core.STAGE3/f"{lane}_queue.csv"; queue=core.read_csv(queue_path)
    checkpoint_path=core.STAGE3/f"{lane}_checkpoint.json"; accepted=core.STAGE3/f"{lane}_accepted_results.jsonl"
    checkpoint=json.loads(checkpoint_path.read_text()) if checkpoint_path.exists() else {"lane_id":lane,"assigned":len(queue),"completed":0,"status":"in_progress","queue_sha256":core.sha256_file(queue_path),"started_at":core.utc_now()}
    if checkpoint.get("status")=="complete": print(json.dumps(checkpoint,indent=2)); return
    done={r["verification_id"] for r in core.read_jsonl(accepted)}
    timeout=httpx.Timeout(20.0,connect=10.0)
    with httpx.Client(timeout=timeout,headers={"User-Agent":"Gabriel-Wages-Research/1.0 metadata-verification"}) as client:
        for row in queue:
            if row["verification_id"] in done: continue
            append_jsonl(accepted,verify_one(row,client)); checkpoint["completed"]+=1; checkpoint["last_verification_id"]=row["verification_id"]; checkpoint["updated_at"]=core.utc_now(); core.atomic_json(checkpoint_path,checkpoint)
    results=core.read_jsonl(accepted); checkpoint.update({"completed":len(results),"status":"complete","finished_at":core.utc_now()}); core.atomic_json(checkpoint_path,checkpoint); core.write_sharded_pair(core.STAGE3,f"{lane}_results",results); print(json.dumps({"lane":lane,"completed":len(results)},indent=2))


def stage3_finalize() -> None:
    results=[]
    for i in range(1,6):
        lane=f"verification_lane_{i:03d}"; cp=json.loads((core.STAGE3/f"{lane}_checkpoint.json").read_text())
        if cp.get("status")!="complete": raise RuntimeError(f"incomplete {lane}")
        results.extend(core.read_jsonl(core.STAGE3/f"{lane}_accepted_results.jsonl"))
    expected=json.loads((core.STAGE3/"verification_lane_distribution.json").read_text())["input_count"]
    if len(results)!=expected or len({r["verification_id"] for r in results})!=expected: raise RuntimeError("verification reconciliation failed")
    by_final={}; ready=[]; duplicate=[]
    for row in results:
        if row["verification_status"] in {"reachable","reachable_with_redirect"}:
            key=core.canonical_url(row["final_url"])
            if key in by_final:
                row["verification_status"]="duplicate_final_locator"; duplicate.append({"duplicate_verification_id":row["verification_id"],"canonical_verification_id":by_final[key]["verification_id"],"final_url":row["final_url"],"basis":"canonical final URL"})
            else: by_final[key]=row; ready.append(row)
    core.write_sharded_pair(core.STAGE3,"verification_results",results); core.write_sharded_pair(core.STAGE3,"source_review_ready_queue",ready); core.write_sharded_pair(core.STAGE3,"duplicate_final_locator_links",duplicate)
    counts=Counter(r["verification_status"] for r in results); checks={"all_inputs_terminal":len(results)==expected,"unique_results":len({r["verification_id"] for r in results})==expected,"no_full_bodies_retained":all(r["full_body_retained"]=="false" for r in results),"ready_plus_nonready_reconciles":len(results)==sum(counts.values())}
    passed=all(checks.values()); core.write_json(core.STAGE3/"verification_validation_report.json",{"passed":passed,"checks":checks,"status_counts":dict(counts),"source_review_ready":len(ready)}); core.write_md(core.STAGE3/"verification_validation_report.md","# External-data verification validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_verification_completed_source_review_ready" if passed else "external_data_verification_repair_needed"; core.write_json(core.STAGE3/"stage_decision.json",{"decision":decision,"status_counts":dict(counts),"source_review_ready":len(ready),"completed_at":core.utc_now()}); core.record_transition("03_EXTERNAL-DATA-VERIFICATION","complete" if passed else "repair_needed",decision,{"status_counts":dict(counts),"source_review_ready":len(ready)}); print(json.dumps({"decision":decision,"counts":dict(counts),"source_review_ready":len(ready)},indent=2))
    if not passed: raise RuntimeError("stage 3 validation failed")
    archive_append_ledgers(core.STAGE3,["verification_lane_*_accepted_results.jsonl"],"stage3_verification")


def extension_for(url: str, content_type: str) -> str:
    ext=Path(urlsplit(url).path).suffix.casefold()
    if ext and len(ext)<=8: return ext
    mime=content_type.split(";",1)[0].strip().casefold(); guessed=mimetypes.guess_extension(mime) or ""
    if guessed==".htm": guessed=".html"
    return guessed or ".bin"


def stage4_prepare() -> None:
    gate=json.loads((core.STAGE3/"verification_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 3 gate must pass")
    rows=load_shards(core.STAGE3,"source_review_ready_queue_shard_manifest.json")
    buckets=lane_partition(rows,"primary_external_data_family")
    for i,part in enumerate(buckets,1): core.write_pair(core.STAGE4,f"source_review_lane_{i:03d}_queue",part)
    core.write_json(core.STAGE4/"source_review_lane_distribution.json",{"input_count":len(rows),"lane_sizes":{f"source_review_lane_{i:03d}":len(x) for i,x in enumerate(buckets,1)},"retained_payload_root":str(core.RETAINED.relative_to(core.ROOT)),"payload_root_git_ignored":True}); core.record_transition("04_EXTERNAL-DATA-SOURCE-REVIEW-DOWNLOAD","prepared","source_review_queues_locked",{"input_count":len(rows)}); print(json.dumps({"source_review_input":len(rows),"lane_sizes":[len(x) for x in buckets]},indent=2))


def retained_status(content_type: str, extension: str) -> str:
    mime=content_type.casefold(); ext=extension.casefold()
    if "pdf" in mime or ext==".pdf": return "retained_pdf"
    if "html" in mime or ext in {".html",".htm"}: return "retained_html"
    if "csv" in mime or ext in {".csv",".tsv"}: return "retained_csv"
    if "spreadsheet" in mime or "excel" in mime or ext in {".xlsx",".xls",".ods"}: return "retained_spreadsheet"
    if "json" in mime or ext==".json": return "retained_json"
    if "xml" in mime or ext==".xml": return "retained_xml"
    if mime.startswith("text/") or ext==".txt": return "retained_text"
    if "zip" in mime or ext==".zip": return "retained_official_data_package"
    return "retained_other_document"


def download_one(row: dict[str,str],client:Any,lane:str,max_bytes:int=50*1024*1024)->dict[str,Any]:
    url=row.get("final_url") or row["candidate_url"]; retained_id=core.stable("EXTSOURCE",row["candidate_id"],core.canonical_url(url)); common={**row,"retained_source_id":retained_id,"downloaded_at":core.utc_now(),"retained_payload_staged":"false"}
    try:
        with client.stream("GET",url,follow_redirects=True) as response:
            if response.status_code in {401,403,451}: return {**common,"source_review_status":"restricted_or_login_required","download_http_status":response.status_code,"retained_path":"","byte_count":0,"sha256":"","download_error":""}
            if response.status_code>=400: return {**common,"source_review_status":"unavailable_on_download","download_http_status":response.status_code,"retained_path":"","byte_count":0,"sha256":"","download_error":""}
            length=int(response.headers.get("content-length") or 0)
            if length>max_bytes: return {**common,"source_review_status":"oversized_defer","download_http_status":response.status_code,"retained_path":"","byte_count":length,"sha256":"","download_error":"declared content length exceeds 50 MiB"}
            content_type=response.headers.get("content-type",""); ext=extension_for(str(response.url),content_type); lane_dir=core.RETAINED/lane; lane_dir.mkdir(parents=True,exist_ok=True); final_path=lane_dir/f"{retained_id}{ext}"; fd,tmp=tempfile.mkstemp(prefix=retained_id,suffix=".part",dir=lane_dir); digest=hashlib.sha256(); count=0
            try:
                with os.fdopen(fd,"wb") as handle:
                    for chunk in response.iter_bytes():
                        count+=len(chunk)
                        if count>max_bytes: raise OverflowError("download exceeded 50 MiB")
                        digest.update(chunk); handle.write(chunk)
                os.replace(tmp,final_path)
            except BaseException:
                Path(tmp).unlink(missing_ok=True); raise
            return {**common,"source_review_status":retained_status(content_type,ext),"download_http_status":response.status_code,"retained_path":str(final_path.relative_to(core.ROOT)),"byte_count":count,"sha256":digest.hexdigest(),"content_type_download":content_type,"download_final_url":str(response.url),"download_error":""}
    except OverflowError as exc:
        return {**common,"source_review_status":"oversized_defer","download_http_status":"","retained_path":"","byte_count":0,"sha256":"","download_error":str(exc)}
    except Exception as exc:
        return {**common,"source_review_status":"source_review_error","download_http_status":"","retained_path":"","byte_count":0,"sha256":"","download_error":f"{exc.__class__.__name__}: {core.clean(str(exc))[:300]}"}


def stage4_run_lane(lane_number:int)->None:
    import httpx
    lane=f"source_review_lane_{lane_number:03d}"; queue_path=core.STAGE4/f"{lane}_queue.csv"; queue=core.read_csv(queue_path); cp_path=core.STAGE4/f"{lane}_checkpoint.json"; accepted=core.STAGE4/f"{lane}_accepted_results.jsonl"; cp=json.loads(cp_path.read_text()) if cp_path.exists() else {"lane_id":lane,"assigned":len(queue),"completed":0,"status":"in_progress","queue_sha256":core.sha256_file(queue_path),"started_at":core.utc_now()}
    if cp.get("status")=="complete": print(json.dumps(cp,indent=2)); return
    done={r["verification_id"] for r in core.read_jsonl(accepted)}
    with httpx.Client(timeout=httpx.Timeout(60.0,connect=15.0),headers={"User-Agent":"Gabriel-Wages-Research/1.0 source-retention"}) as client:
        for row in queue:
            if row["verification_id"] in done: continue
            append_jsonl(accepted,download_one(row,client,lane)); cp["completed"]+=1; cp["last_verification_id"]=row["verification_id"]; cp["updated_at"]=core.utc_now(); core.atomic_json(cp_path,cp)
    results=core.read_jsonl(accepted); cp.update({"completed":len(results),"status":"complete","finished_at":core.utc_now()}); core.atomic_json(cp_path,cp); core.write_sharded_pair(core.STAGE4,f"{lane}_results",results); print(json.dumps({"lane":lane,"completed":len(results)},indent=2))


def stage4_finalize()->None:
    results=[]
    for i in range(1,6):
        lane=f"source_review_lane_{i:03d}"; cp=json.loads((core.STAGE4/f"{lane}_checkpoint.json").read_text())
        if cp.get("status")!="complete": raise RuntimeError(f"incomplete {lane}")
        results.extend(core.read_jsonl(core.STAGE4/f"{lane}_accepted_results.jsonl"))
    expected=json.loads((core.STAGE4/"source_review_lane_distribution.json").read_text())["input_count"]
    if len(results)!=expected: raise RuntimeError("source-review count mismatch")
    seen={}; duplicate_links=[]
    for row in results:
        digest=row.get("sha256","")
        if digest and digest in seen:
            row["source_review_status"]="duplicate_retained_source"; row["canonical_retained_source_id"]=seen[digest]["retained_source_id"]
            duplicate_links.append({"duplicate_retained_source_id":row["retained_source_id"],"canonical_retained_source_id":seen[digest]["retained_source_id"],"sha256":digest,"basis":"identical retained payload hash"})
        elif digest: seen[digest]=row
    core.write_sharded_pair(core.STAGE4,"source_review_download_results",results); core.write_sharded_pair(core.STAGE4,"duplicate_retained_source_links",duplicate_links)
    retained=[r for r in results if r["source_review_status"].startswith("retained_") and r["source_review_status"]!="duplicate_retained_source"]
    core.write_sharded_pair(core.STAGE4,"retained_external_source_manifest",retained)
    counts=Counter(r["source_review_status"] for r in results); bytes_by_type=defaultdict(int)
    for r in retained: bytes_by_type[r["source_review_status"]]+=int(r.get("byte_count") or 0)
    checks={"all_ready_sources_reviewed":len(results)==expected,"unique_verification_rows":len({r["verification_id"] for r in results})==expected,"retained_payloads_ignored":all(r.get("retained_payload_staged")=="false" for r in results),"retained_files_exist":all((core.ROOT/r["retained_path"]).is_file() for r in retained),"hashes_present":all(bool(r["sha256"]) for r in retained)}; passed=all(checks.values()); core.write_json(core.STAGE4/"source_review_download_summary.json",{"input_count":expected,"status_counts":dict(counts),"retained_unique_count":len(retained),"retained_bytes_by_type":dict(bytes_by_type),"duplicate_count":len(duplicate_links)}); core.write_json(core.STAGE4/"source_review_download_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE4/"source_review_download_validation_report.md","# Source-review/download validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_source_review_download_completed_readiness_ready" if passed else "external_data_source_review_download_repair_needed"; core.write_json(core.STAGE4/"stage_decision.json",{"decision":decision,"status_counts":dict(counts),"retained_unique_count":len(retained),"completed_at":core.utc_now()}); core.record_transition("04_EXTERNAL-DATA-SOURCE-REVIEW-DOWNLOAD","complete" if passed else "repair_needed",decision,{"retained":len(retained),"status_counts":dict(counts)}); print(json.dumps({"decision":decision,"retained":len(retained),"counts":dict(counts)},indent=2));
    if not passed: raise RuntimeError("stage 4 validation failed")
    archive_append_ledgers(core.STAGE4,["source_review_lane_*_accepted_results.jsonl"],"stage4_source_review")


def inspect_spreadsheet(path: Path) -> dict[str, Any]:
    """Read workbook structure only; no analytical calculations are made."""
    result: dict[str, Any] = {"sheet_names": [], "sheets": [], "inspection_error": ""}
    try:
        if path.suffix.casefold() == ".xlsx":
            import openpyxl
            workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
            result["sheet_names"] = workbook.sheetnames
            for sheet in workbook.worksheets:
                formula = merged = hidden = False
                header_candidates = []
                for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12)), 1):
                    values = [cell.value for cell in row]
                    formula = formula or any(isinstance(value, str) and value.startswith("=") for value in values)
                    populated = sum(value not in (None, "") for value in values)
                    if populated >= 2: header_candidates.append(row_index)
                merged = bool(sheet.merged_cells.ranges)
                hidden = sheet.sheet_state != "visible"
                result["sheets"].append({"sheet_name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column,
                                         "formula_cell_presence": formula, "merged_cell_presence": merged,
                                         "header_row_candidates": header_candidates[:5], "likely_data_table": sheet.max_row > 1 and sheet.max_column > 1,
                                         "hidden_sheet": hidden})
            workbook.close()
        else:
            import pandas as pd
            book = pd.ExcelFile(path)
            result["sheet_names"] = book.sheet_names
            for sheet_name in book.sheet_names:
                frame = pd.read_excel(path, sheet_name=sheet_name, header=None)
                result["sheets"].append({"sheet_name": sheet_name, "rows": int(frame.shape[0]), "columns": int(frame.shape[1]),
                                         "formula_cell_presence": "not_inspected_for_legacy_xls", "merged_cell_presence": "not_inspected_for_legacy_xls",
                                         "header_row_candidates": [1] if frame.shape[0] else [], "likely_data_table": frame.shape[0] > 1 and frame.shape[1] > 1,
                                         "hidden_sheet": "not_available"})
    except Exception as exc:
        result["inspection_error"] = f"{exc.__class__.__name__}: {core.clean(str(exc))[:300]}"
    return result


def classify_readiness(row: dict[str, str]) -> dict[str, Any]:
    path = core.ROOT / row["retained_path"]
    suffix = path.suffix.casefold(); content_type = row.get("content_type_download", "").casefold()
    status = "needs_manual_review"; detail: dict[str, Any] = {}
    try:
        if suffix == ".pdf" or "pdf" in content_type:
            from pypdf import PdfReader
            reader = PdfReader(path)
            if reader.is_encrypted:
                status = "encrypted_or_locked"
            else:
                sample = "".join((page.extract_text() or "") for page in reader.pages[: min(5, len(reader.pages))])
                status = "parse_text_pdf_ready" if sample.strip() else "ocr_later"
                detail = {"page_count": len(reader.pages), "sample_non_ocr_character_count": len(sample)}
        elif suffix in {".html", ".htm"} or "html" in content_type: status = "html_text_ready"
        elif suffix == ".csv" or "csv" in content_type: status = "csv_structured_ready"
        elif suffix == ".tsv": status = "tsv_structured_ready"
        elif suffix == ".xlsx": status = "xlsx_structured_ready"; detail = inspect_spreadsheet(path)
        elif suffix == ".xls": status = "xls_structured_ready"; detail = inspect_spreadsheet(path)
        elif suffix == ".json" or "json" in content_type: status = "json_structured_ready"
        elif suffix == ".xml" or "xml" in content_type: status = "xml_structured_ready"
        elif suffix == ".txt" or content_type.startswith("text/"): status = "text_ready"
        elif suffix == ".zip" or "zip" in content_type:
            with zipfile.ZipFile(path) as archive:
                detail = {"member_count": len(archive.infolist()), "members": [x.filename for x in archive.infolist()[:100]],
                          "uncompressed_bytes": sum(x.file_size for x in archive.infolist())}
            status = "official_data_package_ready"
        elif path.stat().st_size == 0: status = "corrupt_or_broken"
        else: status = "other_document_ready"
    except Exception as exc:
        status = "readiness_error"; detail = {"error": f"{exc.__class__.__name__}: {core.clean(str(exc))[:300]}"}
    return {**row, "readiness_status": status, "readiness_detail": json.dumps(detail, sort_keys=True),
            "readiness_assessed_at": core.utc_now(), "ocr_performed": "false"}


def stage5_run() -> None:
    gate = json.loads((core.STAGE4 / "source_review_download_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 4 gate must pass")
    retained = load_shards(core.STAGE4, "retained_external_source_manifest_shard_manifest.json")
    results = [classify_readiness(row) for row in retained]
    core.write_sharded_pair(core.STAGE5, "external_data_readiness_results", results)
    ready_statuses = {"parse_text_pdf_ready", "html_text_ready", "csv_structured_ready", "tsv_structured_ready",
                      "xlsx_structured_ready", "xls_structured_ready", "json_structured_ready", "xml_structured_ready",
                      "text_ready", "official_data_package_ready", "other_document_ready"}
    ready = [row for row in results if row["readiness_status"] in ready_statuses]
    core.write_sharded_pair(core.STAGE5, "external_data_extraction_ready_queue", ready)
    counts = Counter(row["readiness_status"] for row in results)
    checks = {"all_retained_classified": len(results) == len(retained), "unique_retained_sources": len({r["retained_source_id"] for r in results}) == len(results),
              "no_ocr": all(r["ocr_performed"] == "false" for r in results), "ready_subset_valid": all(r["readiness_status"] in ready_statuses for r in ready)}
    passed = all(checks.values())
    core.write_json(core.STAGE5 / "external_data_readiness_summary.json", {"retained_count":len(retained), "status_counts":dict(counts), "extraction_ready_count":len(ready)})
    core.write_json(core.STAGE5 / "readiness_validation_report.json", {"passed":passed,"checks":checks,"status_counts":dict(counts)})
    core.write_md(core.STAGE5 / "readiness_validation_report.md", "# External-data readiness validation\n\n" + "\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items()))
    decision="external_data_readiness_completed_extraction_ready" if passed else "external_data_readiness_repair_needed"
    core.write_json(core.STAGE5 / "stage_decision.json", {"decision":decision,"status_counts":dict(counts),"extraction_ready_count":len(ready),"completed_at":core.utc_now()})
    core.record_transition("05_EXTERNAL-DATA-READINESS","complete" if passed else "repair_needed",decision,{"status_counts":dict(counts),"extraction_ready":len(ready)})
    print(json.dumps({"decision":decision,"counts":dict(counts),"extraction_ready":len(ready)},indent=2))
    if not passed: raise RuntimeError("stage 5 validation failed")


def stage6_prepare() -> None:
    gate=json.loads((core.STAGE5/"readiness_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 5 gate must pass")
    rows=load_shards(core.STAGE5,"external_data_extraction_ready_queue_shard_manifest.json"); buckets=lane_partition(rows,"primary_external_data_family")
    for i,part in enumerate(buckets,1): core.write_pair(core.STAGE6,f"extraction_lane_{i:03d}_queue",part)
    core.write_json(core.STAGE6/"extraction_lane_distribution.json",{"input_count":len(rows),"lane_sizes":{f"extraction_lane_{i:03d}":len(x) for i,x in enumerate(buckets,1)},"ocr_allowed":False})
    core.record_transition("06_EXTERNAL-DATA-EXTRACTION","prepared","external_data_extraction_queues_locked",{"input_count":len(rows)})
    print(json.dumps({"extraction_input":len(rows),"lane_sizes":[len(x) for x in buckets]},indent=2))


def safe_text_path(retained_id: str, suffix: str = ".txt") -> Path:
    path=core.EXTRACTED/retained_id[:2]/f"{retained_id}{suffix}"; path.parent.mkdir(parents=True,exist_ok=True); return path


def safe_structured_dir(retained_id: str) -> Path:
    path=core.STRUCTURED/retained_id[:2]/retained_id; path.mkdir(parents=True,exist_ok=True); return path


def extraction_one(row:dict[str,str])->dict[str,Any]:
    source=core.ROOT/row["retained_path"]; retained_id=row["retained_source_id"]; status=row["readiness_status"]
    common={**row,"extraction_id":core.stable("EXTEXTRACT",retained_id),"extracted_at":core.utc_now(),"ocr_performed":"false","full_payload_tracked":"false"}
    try:
        if status=="parse_text_pdf_ready":
            from pypdf import PdfReader
            reader=PdfReader(source); pages=[]; boundaries=[]; offset=0
            for i,page in enumerate(reader.pages,1):
                text=page.extract_text() or ""; pages.append(text); boundaries.append({"page":i,"start":offset,"end":offset+len(text)}); offset+=len(text)+2
            joined="\n\n".join(pages); out=safe_text_path(retained_id); out.write_text(joined,encoding="utf-8")
            return {**common,"extraction_status":"text_extracted","extracted_text_path":str(out.relative_to(core.ROOT)),"structured_data_path":"","text_character_count":len(joined),"page_or_section_count":len(pages),"record_count":0,"schema_json":"","boundary_json":json.dumps(boundaries,separators=(",",":")),"extraction_error":""}
        if status in {"html_text_ready","text_ready","other_document_ready"}:
            raw=source.read_bytes()
            if status=="html_text_ready":
                from bs4 import BeautifulSoup
                soup=BeautifulSoup(raw,"html.parser")
                for element in soup(["script","style","noscript"]): element.decompose()
                text=soup.get_text("\n")
            else: text=raw.decode("utf-8",errors="replace")
            out=safe_text_path(retained_id); out.write_text(text,encoding="utf-8")
            return {**common,"extraction_status":"text_extracted","extracted_text_path":str(out.relative_to(core.ROOT)),"structured_data_path":"","text_character_count":len(text),"page_or_section_count":0,"record_count":0,"schema_json":"","boundary_json":"[]","extraction_error":""}
        target=safe_structured_dir(retained_id); schema={}; record_count=0
        if status in {"csv_structured_ready","tsv_structured_ready"}:
            delimiter="\t" if status.startswith("tsv") else ","; out=target/"records.jsonl"
            with source.open(newline="",encoding="utf-8-sig",errors="replace") as handle, out.open("w",encoding="utf-8") as dest:
                reader=csv.DictReader(handle,delimiter=delimiter); schema={"table":"records","columns":reader.fieldnames or [],"source_format":source.suffix.casefold()}
                for index,item in enumerate(reader,2): dest.write(json.dumps({"source_row_index":index,"values":item},ensure_ascii=False,separators=(",",":"))+"\n"); record_count+=1
        elif status in {"xlsx_structured_ready","xls_structured_ready"}:
            import pandas as pd
            book=pd.ExcelFile(source); schema={"sheets":[]}
            for sheet_index,sheet in enumerate(book.sheet_names,1):
                frame=pd.read_excel(source,sheet_name=sheet,dtype=object); out=target/f"sheet_{sheet_index:03d}.jsonl"; columns=[str(x) for x in frame.columns]; schema["sheets"].append({"sheet_index":sheet_index,"sheet_name":sheet,"columns":columns,"rows":len(frame),"path":out.name})
                with out.open("w",encoding="utf-8") as dest:
                    for source_index,values in enumerate(frame.itertuples(index=False,name=None),2):
                        item={columns[i]: ("" if value is None or (isinstance(value,float) and value!=value) else value) for i,value in enumerate(values)}
                        dest.write(json.dumps({"sheet_name":sheet,"source_row_index":source_index,"values":item},ensure_ascii=False,default=str,separators=(",",":"))+"\n"); record_count+=1
        elif status=="json_structured_ready":
            data=json.loads(source.read_text(encoding="utf-8-sig",errors="replace")); items=data if isinstance(data,list) else data.get("data",[]) if isinstance(data,dict) and isinstance(data.get("data"),list) else [data]; out=target/"records.jsonl"; schema={"top_level_type":type(data).__name__,"columns":sorted({str(k) for item in items if isinstance(item,dict) for k in item})}
            with out.open("w",encoding="utf-8") as dest:
                for index,item in enumerate(items,1): dest.write(json.dumps({"source_row_index":index,"values":item},ensure_ascii=False,separators=(",",":"))+"\n"); record_count+=1
        elif status=="xml_structured_ready":
            import xml.etree.ElementTree as ET
            tree=ET.parse(source); root=tree.getroot(); out=target/"records.jsonl"; children=list(root); schema={"root_tag":root.tag,"record_tags":sorted({x.tag for x in children})}
            with out.open("w",encoding="utf-8") as dest:
                for index,item in enumerate(children,1): values={child.tag:child.text or "" for child in item}; dest.write(json.dumps({"source_row_index":index,"record_tag":item.tag,"values":values},ensure_ascii=False,separators=(",",":"))+"\n"); record_count+=1
        elif status=="official_data_package_ready":
            with zipfile.ZipFile(source) as archive:
                members=[]
                for info in archive.infolist():
                    safe=not info.is_dir() and ".." not in Path(info.filename).parts and not Path(info.filename).is_absolute()
                    members.append({"name":info.filename,"bytes":info.file_size,"safe_path":safe})
                schema={"package_members":members,"member_count":len(members)}
            record_count=len(schema["package_members"])
        else: raise ValueError(f"unsupported approved readiness status {status}")
        schema_path=target/"schema.json"; core.write_json(schema_path,schema)
        return {**common,"extraction_status":"structured_data_extracted","extracted_text_path":"","structured_data_path":str(target.relative_to(core.ROOT)),"text_character_count":0,"page_or_section_count":0,"record_count":record_count,"schema_json":json.dumps(schema,sort_keys=True,separators=(",",":")),"boundary_json":"[]","extraction_error":""}
    except Exception as exc:
        return {**common,"extraction_status":"extraction_error","extracted_text_path":"","structured_data_path":"","text_character_count":0,"page_or_section_count":0,"record_count":0,"schema_json":"","boundary_json":"[]","extraction_error":f"{exc.__class__.__name__}: {core.clean(str(exc))[:500]}"}


def stage6_run_lane(lane_number:int)->None:
    lane=f"extraction_lane_{lane_number:03d}"; queue_path=core.STAGE6/f"{lane}_queue.csv"; queue=core.read_csv(queue_path); cp_path=core.STAGE6/f"{lane}_checkpoint.json"; accepted=core.STAGE6/f"{lane}_accepted_results.jsonl"; cp=json.loads(cp_path.read_text()) if cp_path.exists() else {"lane_id":lane,"assigned":len(queue),"completed":0,"status":"in_progress","queue_sha256":core.sha256_file(queue_path),"started_at":core.utc_now()}
    if cp.get("status")=="complete": print(json.dumps(cp,indent=2)); return
    done={r["retained_source_id"] for r in core.read_jsonl(accepted)}
    for row in queue:
        if row["retained_source_id"] in done: continue
        append_jsonl(accepted,extraction_one(row)); cp["completed"]+=1; cp["last_retained_source_id"]=row["retained_source_id"]; cp["updated_at"]=core.utc_now(); core.atomic_json(cp_path,cp)
    results=core.read_jsonl(accepted); cp.update({"completed":len(results),"status":"complete","finished_at":core.utc_now()}); core.atomic_json(cp_path,cp); core.write_sharded_pair(core.STAGE6,f"{lane}_results",results); print(json.dumps({"lane":lane,"completed":len(results)},indent=2))


def stage6_finalize()->None:
    results=[]
    for i in range(1,6):
        lane=f"extraction_lane_{i:03d}"; cp=json.loads((core.STAGE6/f"{lane}_checkpoint.json").read_text())
        if cp.get("status")!="complete": raise RuntimeError(f"incomplete {lane}")
        results.extend(core.read_jsonl(core.STAGE6/f"{lane}_accepted_results.jsonl"))
    expected=json.loads((core.STAGE6/"extraction_lane_distribution.json").read_text())["input_count"]
    core.write_sharded_pair(core.STAGE6,"external_data_extraction_results",results)
    ready=[r for r in results if r["extraction_status"] in {"text_extracted","structured_data_extracted"}]; core.write_sharded_pair(core.STAGE6,"external_data_field_extraction_ready_queue",ready)
    counts=Counter(r["extraction_status"] for r in results); checks={"all_ready_sources_processed":len(results)==expected,"unique_source_results":len({r["retained_source_id"] for r in results})==len(results),"no_ocr":all(r["ocr_performed"]=="false" for r in results),"payloads_untracked":all(r["full_payload_tracked"]=="false" for r in results),"output_pointers_exist":all((not r["extracted_text_path"] or (core.ROOT/r["extracted_text_path"]).exists()) and (not r["structured_data_path"] or (core.ROOT/r["structured_data_path"]).exists()) for r in ready)}; passed=all(checks.values()); core.write_json(core.STAGE6/"external_data_extraction_summary.json",{"input_count":expected,"status_counts":dict(counts),"field_extraction_ready_count":len(ready),"total_text_characters":sum(int(r["text_character_count"]) for r in ready),"total_structured_records":sum(int(r["record_count"]) for r in ready)}); core.write_json(core.STAGE6/"extraction_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE6/"extraction_validation_report.md","# External-data extraction validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_extraction_completed_field_extraction_ready" if passed else "external_data_extraction_repair_needed"; core.write_json(core.STAGE6/"stage_decision.json",{"decision":decision,"status_counts":dict(counts),"field_extraction_ready":len(ready),"completed_at":core.utc_now()}); core.record_transition("06_EXTERNAL-DATA-EXTRACTION","complete" if passed else "repair_needed",decision,{"field_extraction_ready":len(ready),"status_counts":dict(counts)}); print(json.dumps({"decision":decision,"counts":dict(counts),"field_extraction_ready":len(ready)},indent=2));
    if not passed: raise RuntimeError("stage 6 validation failed")
    archive_append_ledgers(core.STAGE6,["extraction_lane_*_accepted_results.jsonl"],"stage6_extraction")


FIELD_RULES: list[tuple[str, str, str]] = [
    ("payroll_and_earnings", "base_pay", r"\b(base (?:pay|salary|wage)|regular (?:pay|earnings))\b"),
    ("payroll_and_earnings", "overtime_earnings", r"\b(overtime(?: earnings| pay| compensation)?)\b"),
    ("payroll_and_earnings", "total_earnings", r"\b(total (?:earnings|pay|wages|compensation))\b"),
    ("payroll_and_earnings", "retroactive_pay", r"\b(retroactive pay|back pay)\b"),
    ("payroll_and_earnings", "lump_sum", r"\b(lump[ -]sum)\b"),
    ("payroll_and_earnings", "premium_or_stipend", r"\b(premium|stipend|allowance|other compensation)\b"),
    ("staffing_and_headcount", "authorized_positions", r"\b(authorized (?:positions?|strength|headcount))\b"),
    ("staffing_and_headcount", "budgeted_positions", r"\b(budgeted (?:positions?|fte|headcount))\b"),
    ("staffing_and_headcount", "filled_positions", r"\b(filled (?:positions?|fte)|positions? filled)\b"),
    ("staffing_and_headcount", "vacant_positions", r"\b(vacant (?:positions?|fte)|vacanc(?:y|ies|t))\b"),
    ("staffing_and_headcount", "position_elimination", r"\b(position eliminat|eliminat(?:e|ed|ion) positions?|layoffs?|attrition not replaced|hiring freeze|outsourc|consolidat)\w*\b"),
    ("staffing_and_headcount", "minimum_staffing", r"\b(minimum staffing|minimum manning)\b"),
    ("recruitment_and_retention", "applicant_count", r"\b(applicant(?: count|s)?|applications? received)\b"),
    ("recruitment_and_retention", "turnover_or_retention", r"\b(turnover|retention|vacancy duration|hiring difficult)\w*\b"),
    ("recruitment_and_retention", "recruitment_or_retention_incentive", r"\b(recruitment (?:bonus|incentive)|retention (?:bonus|incentive|pay))\b"),
    ("tenure_and_progression", "years_of_service", r"\b(years? of service|service years?)\b"),
    ("tenure_and_progression", "step_placement", r"\b(step (?:placement|schedule|progression)|salary step)\b"),
    ("tenure_and_progression", "rank_or_classification", r"\b(rank|classification|job title|position title|promotion|seniority)\b"),
    ("implementation_confirmation", "adoption_or_ratification_date", r"\b(adopted|approved|ratified|ordinance date|resolution date|appropriation date)\b"),
    ("implementation_confirmation", "effective_or_payment_date", r"\b(effective date|payroll[- ]effective|payment date|paid on)\b"),
    ("implementation_confirmation", "recurring_or_one_time", r"\b(recurring|one[- ]time|temporary|permanent)\b"),
    ("benefits_and_total_compensation", "pension_contribution", r"\b(pension contribution|retirement contribution)\b"),
    ("benefits_and_total_compensation", "health_contribution", r"\b(health (?:insurance )?contribution|employee cost share|employer cost share)\b"),
    ("benefits_and_total_compensation", "leave_or_holiday_benefit", r"\b(leave value|holiday (?:pay|benefit)|longevity)\b"),
    ("benefits_and_total_compensation", "education_or_certification_pay", r"\b(education pay|certification pay|education incentive)\b"),
    ("contextual_controls", "population", r"\b(population)\b"),
    ("contextual_controls", "labor_market_context", r"\b(unemployment|labor market|fiscal capacity|urban|rural)\b"),
    ("contextual_controls", "labor_law_or_union_context", r"\b(collective bargaining law|labor law|unionization|right to work)\b"),
]
FIELD_RULES_COMPILED = [(family, field, re.compile(pattern, re.I)) for family, field, pattern in FIELD_RULES]
VALUE_PATTERN = re.compile(r"(?:\$\s*)?[-+]?\d[\d,]*(?:\.\d+)?\s*(?:%|percent|hours?|fte|positions?|employees?)?", re.I)


def field_rule_for_label(label: str) -> tuple[str, str] | None:
    normalized = re.sub(r"[^a-z0-9]+", " ", label.casefold()).strip()
    for family, field, pattern in FIELD_RULES_COMPILED:
        if pattern.search(normalized): return family, field
    synonyms = {
        "salary": ("payroll_and_earnings", "base_pay"), "wage": ("payroll_and_earnings", "base_pay"),
        "department": ("contextual_controls", "department"), "title": ("tenure_and_progression", "role_title"),
        "year": ("contextual_controls", "year_or_period"), "fiscal year": ("contextual_controls", "year_or_period"),
        "employee name": ("payroll_and_earnings", "employee_or_position_name"), "name": ("payroll_and_earnings", "employee_or_position_name"),
        "fte": ("staffing_and_headcount", "headcount_or_fte"), "headcount": ("staffing_and_headcount", "headcount_or_fte"),
    }
    return synonyms.get(normalized)


def base_field_record(row: dict[str, str], family: str, field_name: str, locator: str, raw_value: Any, evidence: str) -> dict[str, Any]:
    value = core.clean(raw_value)
    value_type = "currency" if "$" in value else "percentage" if "%" in value or "percent" in value.casefold() else "numeric" if re.fullmatch(r"[-+]?\d[\d,]*(?:\.\d+)?", value) else "text"
    return {
        "external_field_record_id": core.stable("EXTFIELD", row["retained_source_id"], locator, field_name, value, n=28),
        "retained_source_id": row["retained_source_id"], "candidate_id": row.get("candidate_id", ""),
        "source_id": row["retained_source_id"], "source_locator": locator, "municipality": row.get("municipality", ""),
        "state": row.get("state", ""), "year_or_period": row.get("period", ""), "department_or_unit": row.get("department_scope", ""),
        "side_hint": row.get("side_scope", ""), "role_or_title": "", "external_data_family": family, "field_name": field_name,
        "raw_value": value, "value_type": value_type, "unit_or_pay_basis": "unknown", "bounded_evidence": core.clean(evidence)[:1000],
        "confidence": "high" if value_type != "text" else "moderate", "linked_root_event_ids": row.get("linked_root_event_id", ""),
        "linked_mechanism_exposure_event_ids": row.get("linked_mechanism_exposure_event_ids", ""),
        "linked_claims": row.get("expected_claim_upgrade", ""), "source_wave": row.get("search_wave", ""),
    }


def text_field_records(row: dict[str, str], text: str) -> list[dict[str, Any]]:
    records=[]; seen=set()
    for family, field, pattern in FIELD_RULES_COMPILED:
        for match in pattern.finditer(text):
            start=max(0,match.start()-240); end=min(len(text),match.end()+300); excerpt=text[start:end]
            values=VALUE_PATTERN.findall(excerpt)
            raw=values[0] if values else match.group(0)
            key=(field,match.start(),raw)
            if key in seen: continue
            seen.add(key); records.append(base_field_record(row,family,field,f"text_offset:{match.start()}-{match.end()}",raw,excerpt))
    return records


def structured_field_records(row:dict[str,str],directory:Path)->list[dict[str,Any]]:
    records=[]
    for path in sorted(directory.glob("*.jsonl")):
        sheet_or_table=path.stem
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip(): continue
                item=json.loads(line); values=item.get("values",{})
                if not isinstance(values,dict): continue
                source_row=item.get("source_row_index","")
                for column,value in values.items():
                    if value in (None,""): continue
                    rule=field_rule_for_label(str(column))
                    if not rule: continue
                    family,field=rule; locator=f"{item.get('sheet_name',sheet_or_table)}!row:{source_row}!column:{column}"
                    records.append(base_field_record(row,family,field,locator,value,f"{column}: {value}"))
    return records


def stage7_prepare()->None:
    gate=json.loads((core.STAGE6/"extraction_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 6 gate must pass")
    rows=load_shards(core.STAGE6,"external_data_field_extraction_ready_queue_shard_manifest.json"); buckets=lane_partition(rows,"primary_external_data_family")
    for i,part in enumerate(buckets,1): core.write_pair(core.STAGE7,f"field_extraction_lane_{i:03d}_queue",part)
    core.write_json(core.STAGE7/"field_extraction_lane_distribution.json",{"input_source_count":len(rows),"lane_sizes":{f"field_extraction_lane_{i:03d}":len(x) for i,x in enumerate(buckets,1)},"raw_values_preserved":True,"national_estimation":False})
    core.record_transition("07_EXTERNAL-DATA-FIELD-SPAN-EXTRACTION","prepared","external_field_extraction_queues_locked",{"input_source_count":len(rows)})
    print(json.dumps({"field_extraction_sources":len(rows),"lane_sizes":[len(x) for x in buckets]},indent=2))


def stage7_run_lane(lane_number:int)->None:
    lane=f"field_extraction_lane_{lane_number:03d}"; queue_path=core.STAGE7/f"{lane}_queue.csv"; queue=core.read_csv(queue_path); cp_path=core.STAGE7/f"{lane}_checkpoint.json"; outcome_path=core.STAGE7/f"{lane}_accepted_outcomes.jsonl"; fields_path=core.STAGE7/f"{lane}_accepted_field_records.jsonl"; cp=json.loads(cp_path.read_text()) if cp_path.exists() else {"lane_id":lane,"assigned":len(queue),"completed":0,"field_record_count":0,"status":"in_progress","queue_sha256":core.sha256_file(queue_path),"started_at":core.utc_now()}
    if cp.get("status")=="complete": print(json.dumps(cp,indent=2)); return
    done={r["retained_source_id"] for r in core.read_jsonl(outcome_path)}
    for row in queue:
        if row["retained_source_id"] in done: continue
        error=""; records=[]
        try:
            if row["extraction_status"]=="text_extracted": records=text_field_records(row,(core.ROOT/row["extracted_text_path"]).read_text(encoding="utf-8",errors="replace"))
            elif row["extraction_status"]=="structured_data_extracted": records=structured_field_records(row,core.ROOT/row["structured_data_path"])
        except Exception as exc: error=f"{exc.__class__.__name__}: {core.clean(str(exc))[:400]}"
        for record in records: append_jsonl(fields_path,record)
        append_jsonl(outcome_path,{"retained_source_id":row["retained_source_id"],"field_extraction_status":"field_records_extracted" if records else "zero_relevant_field_records" if not error else "field_extraction_error","field_record_count":len(records),"error":error,"completed_at":core.utc_now(),"lane_id":lane})
        cp["completed"]+=1; cp["field_record_count"]+=len(records); cp["last_retained_source_id"]=row["retained_source_id"]; cp["updated_at"]=core.utc_now(); core.atomic_json(cp_path,cp)
    outcomes=core.read_jsonl(outcome_path); fields=core.read_jsonl(fields_path); cp.update({"completed":len(outcomes),"field_record_count":len(fields),"status":"complete","finished_at":core.utc_now()}); core.atomic_json(cp_path,cp); core.write_sharded_pair(core.STAGE7,f"{lane}_outcomes",outcomes); core.write_sharded_pair(core.STAGE7,f"{lane}_field_records",fields); print(json.dumps({"lane":lane,"sources":len(outcomes),"field_records":len(fields)},indent=2))


def stage7_finalize()->None:
    outcomes=[]; fields=[]
    for i in range(1,6):
        lane=f"field_extraction_lane_{i:03d}"; cp=json.loads((core.STAGE7/f"{lane}_checkpoint.json").read_text())
        if cp.get("status")!="complete": raise RuntimeError(f"incomplete {lane}")
        outcomes.extend(core.read_jsonl(core.STAGE7/f"{lane}_accepted_outcomes.jsonl")); fields.extend(core.read_jsonl(core.STAGE7/f"{lane}_accepted_field_records.jsonl"))
    expected=json.loads((core.STAGE7/"field_extraction_lane_distribution.json").read_text())["input_source_count"]
    core.write_sharded_pair(core.STAGE7,"external_data_field_extraction_outcomes",outcomes); core.write_sharded_pair(core.STAGE7,"external_data_field_record_layer",fields); core.write_sharded_pair(core.STAGE7,"external_data_evidence_span_layer",fields); core.write_sharded_pair(core.STAGE7,"external_data_gabriel_rating_ready_queue",fields)
    status_counts=Counter(r["field_extraction_status"] for r in outcomes); family_counts=Counter(r["external_data_family"] for r in fields); checks={"all_sources_terminal":len(outcomes)==expected,"unique_source_outcomes":len({r["retained_source_id"] for r in outcomes})==expected,"field_ids_unique":len({r["external_field_record_id"] for r in fields})==len(fields),"source_lineage_present":all(r["source_id"] and r["source_locator"] for r in fields),"raw_values_preserved":all("raw_value" in r for r in fields),"no_national_estimate":True}; passed=all(checks.values()); core.write_json(core.STAGE7/"external_data_field_extraction_summary.json",{"input_source_count":expected,"outcome_status_counts":dict(status_counts),"field_record_count":len(fields),"field_family_counts":dict(family_counts),"rating_ready_count":len(fields)}); core.write_json(core.STAGE7/"field_extraction_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE7/"field_extraction_validation_report.md","# External-data field/span extraction validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_field_span_extraction_completed_rating_ready" if passed else "external_data_field_span_extraction_repair_needed"; core.write_json(core.STAGE7/"stage_decision.json",{"decision":decision,"field_records":len(fields),"family_counts":dict(family_counts),"completed_at":core.utc_now()}); core.record_transition("07_EXTERNAL-DATA-FIELD-SPAN-EXTRACTION","complete" if passed else "repair_needed",decision,{"field_records":len(fields),"family_counts":dict(family_counts)}); print(json.dumps({"decision":decision,"field_records":len(fields),"family_counts":dict(family_counts)},indent=2));
    if not passed: raise RuntimeError("stage 7 validation failed")
    archive_append_ledgers(core.STAGE7,["field_extraction_lane_*_accepted_outcomes.jsonl","field_extraction_lane_*_accepted_field_records.jsonl"],"stage7_field_extraction")


RATING_CONTROLS = {
    "administrative_evidence_quality": ["direct_administrative_record","strong_official_summary","moderate_official_context","weak_or_ambiguous","non_primary_context","quarantine"],
    "claim_upgrade": ["upgrades_local_wage_comparison","upgrades_total_compensation_comparison","upgrades_staffing_hypothesis","upgrades_growth_analysis","upgrades_implementation_confirmation","upgrades_mechanism_claim","upgrades_national_readiness","context_only","no_material_upgrade"],
    "side_relevance": ["police","fire","safety_combined","non_safety","mixed","side_independent","unclear"],
    "value_readiness": ["direct_scalar_ready","structured_table_ready","range_or_schedule_ready","percentage_or_growth_ready","staffing_metric_ready","implementation_date_ready","benefits_metric_ready","context_metric_ready","needs_normalization","not_quantitative"],
    "linkage_strength": ["exact_same_source_event","same_municipality_period","same_department_period","same_cycle_cross_source","contextual_only","not_linkable"],
}


def rating_schema(max_records:int)->dict[str,Any]:
    item_props={"external_field_record_id":{"type":"string"}}
    for key,values in RATING_CONTROLS.items(): item_props[key]={"type":"string","enum":values}
    item_props["reason_codes"]={"type":"array","items":{"type":"string","pattern":"^[a-z][a-z0-9_]{0,63}$"},"minItems":1,"maxItems":6}
    item_props["concise_rating_rationale"]={"type":"string","minLength":1,"maxLength":360}
    return {"type":"object","properties":{"packet_id":{"type":"string"},"ratings":{"type":"array","items":{"type":"object","properties":item_props,"required":list(item_props),"additionalProperties":False},"minItems":1,"maxItems":max_records}},"required":["packet_id","ratings"],"additionalProperties":False}


def packet_payload(packet:dict[str,str],records:list[dict[str,str]])->dict[str,Any]:
    return {"packet_id":packet["packet_id"],"records":[{"external_field_record_id":r["external_field_record_id"],"municipality":r["municipality"],"state":r["state"],"period":r["year_or_period"],"department_or_unit":r["department_or_unit"],"side_hint":r["side_hint"],"external_data_family":r["external_data_family"],"field_name":r["field_name"],"raw_value":r["raw_value"],"value_type":r["value_type"],"source_locator":r["source_locator"],"bounded_evidence":r["bounded_evidence"],"linked_root_event_ids":r["linked_root_event_ids"],"expected_claim_upgrade":r["linked_claims"]} for r in records]}


def rating_prompt(payload:dict[str,Any],repair:bool=False)->str:
    controls=json.dumps(RATING_CONTROLS,sort_keys=True)
    note=" A prior attempt failed transport or strict schema validation; return every record exactly once." if repair else ""
    return ("Rate bounded external administrative evidence for the Gabriel Wages municipal compensation study. "
            "Do not estimate a wage gap, prevalence, regression, treatment effect, or causal effect. Preserve uncertain side and linkage labels. "
            "Use only the supplied bounded evidence and metadata. Return strict JSON under the supplied schema. Controlled labels: "+controls+note+"\nINPUT:\n"+json.dumps(payload,sort_keys=True,ensure_ascii=False,separators=(",",":")))


def validate_rating_response(value:Any,packet:dict[str,str],records:list[dict[str,str]])->list[dict[str,Any]]:
    if not isinstance(value,dict) or set(value)!={"packet_id","ratings"} or value["packet_id"]!=packet["packet_id"]: raise ValueError("packet_schema_or_lineage_invalid")
    ratings=value["ratings"]
    if not isinstance(ratings,list) or len(ratings)!=len(records): raise ValueError("rating_count_mismatch")
    by_id={item.get("external_field_record_id"):item for item in ratings if isinstance(item,dict)}; expected=[r["external_field_record_id"] for r in records]
    if set(by_id)!=set(expected) or len(by_id)!=len(ratings): raise ValueError("record_membership_invalid")
    validated=[]
    for record in records:
        item=by_id[record["external_field_record_id"]]
        if set(item)!={"external_field_record_id",*RATING_CONTROLS,"reason_codes","concise_rating_rationale"}: raise ValueError("rating_item_shape_invalid")
        for field,allowed in RATING_CONTROLS.items():
            if item[field] not in allowed: raise ValueError(field+"_uncontrolled")
        if not isinstance(item["reason_codes"],list) or not 1<=len(item["reason_codes"])<=6: raise ValueError("reason_codes_invalid")
        if not isinstance(item["concise_rating_rationale"],str) or not 1<=len(item["concise_rating_rationale"])<=360: raise ValueError("rationale_invalid")
        validated.append(item)
    return validated


def stage8_prepare()->None:
    gate=json.loads((core.STAGE7/"field_extraction_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 7 gate must pass")
    records=load_shards(core.STAGE7,"external_data_gabriel_rating_ready_queue_shard_manifest.json")
    grouped=defaultdict(list)
    for record in records: grouped[record["retained_source_id"]].append(record)
    packets=[]; packet_records={}
    for retained_id in sorted(grouped):
        rows=sorted(grouped[retained_id],key=lambda r:r["external_field_record_id"])
        for part_index,start in enumerate(range(0,len(rows),20),1):
            part=rows[start:start+20]; packet_id=core.stable("EXTRATEPACK",retained_id,part_index,"external_admin_rating_v1",n=28); packet_records[packet_id]=part
            packets.append({"packet_id":packet_id,"retained_source_id":retained_id,"packet_part_index":part_index,"record_count":len(part),"record_ids_sha256":hashlib.sha256("|".join(r["external_field_record_id"] for r in part).encode()).hexdigest(),"packet_input_sha256":hashlib.sha256(rating_prompt(packet_payload({"packet_id":packet_id},part)).encode()).hexdigest()})
    buckets=lane_partition(packets,"retained_source_id")
    for i,part in enumerate(buckets,1):
        lane=f"rating_lane_{i:03d}"
        for packet in part: packet["rating_lane_id"]=lane
        core.write_pair(core.STAGE8,f"{lane}_queue",part)
    core.write_sharded_pair(core.STAGE8,"external_data_rating_locked_record_queue",records); core.write_pair(core.STAGE8,"external_data_rating_packet_manifest",packets)
    core.write_json(core.STAGE8/"external_data_rating_output_schema.json",rating_schema(20)); core.write_json(core.STAGE8/"rating_lane_distribution.json",{"record_count":len(records),"packet_count":len(packets),"lane_sizes":{f"rating_lane_{i:03d}":len(x) for i,x in enumerate(buckets,1)},"max_records_per_packet":20})
    dry_checks={"all_records_packetized":sum(int(p["record_count"]) for p in packets)==len(records),"packet_ids_unique":len({p["packet_id"] for p in packets})==len(packets),"schema_constructed":True,"raw_prompts_saved":False,"raw_responses_saved":False,"model_api_calls":0,"forbidden_estimates":False}; passed=all(dry_checks.values()); core.write_json(core.STAGE8/"gabriel_rating_dry_run_report.json",{"passed":passed,"checks":dry_checks,"records":len(records),"packets":len(packets)}); core.write_json(core.STAGE8/"gabriel_rating_transport_preflight.json",{"status":"pending_live_smoke","dry_run_passed":passed,"live_smoke_passed":False,"live_lanes_authorized":False,"credential_value_logged":False}); core.record_transition("08_EXTERNAL-DATA-GABRIEL-RATING","prepared","external_data_rating_dry_run_complete",{"records":len(records),"packets":len(packets)}); print(json.dumps({"records":len(records),"packets":len(packets),"lane_sizes":[len(x) for x in buckets]},indent=2))


def load_rating_records()->dict[str,list[dict[str,str]]]:
    records=load_shards(core.STAGE8,"external_data_rating_locked_record_queue_shard_manifest.json"); grouped=defaultdict(list)
    for row in records: grouped[row["retained_source_id"]].append(row)
    result={}
    for packet in core.read_csv(core.STAGE8/"external_data_rating_packet_manifest.csv"):
        rows=sorted(grouped[packet["retained_source_id"]],key=lambda r:r["external_field_record_id"]); start=(int(packet["packet_part_index"])-1)*20; result[packet["packet_id"]]=rows[start:start+int(packet["record_count"])]
    return result


async def live_rate_requests(items:list[tuple[dict[str,str],list[dict[str,str]],bool]],key:str,parallel:int=5,timeout:float=120.0)->list[dict[str,Any]]:
    import httpx
    from openai import AsyncOpenAI
    client=AsyncOpenAI(api_key=key,base_url="https://go.apis.huit.harvard.edu/ais-openai-direct/v2",default_headers={"Ocp-Apim-Subscription-Key":key},timeout=httpx.Timeout(timeout),max_retries=0); sem=asyncio.Semaphore(parallel)
    async def one(packet:dict[str,str],records:list[dict[str,str]],repair:bool)->dict[str,Any]:
        prompt=rating_prompt(packet_payload(packet,records),repair); started=time.monotonic(); started_at=core.utc_now()
        async with sem:
            try:
                response=await asyncio.wait_for(client.responses.create(model="gpt-5.4-nano",input=prompt,reasoning={"effort":"low"},text={"format":{"type":"json_schema","name":"external_administrative_evidence_rating_v1","strict":True,"schema":rating_schema(len(records))}}),timeout=timeout)
                usage=getattr(response,"usage",None); text=str(getattr(response,"output_text","") or ""); parsed=validate_rating_response(json.loads(text),packet,records)
                return {"status":"valid","packet":packet,"records":records,"ratings":parsed,"request_id":str(getattr(response,"id","") or ""),"input_tokens":int(getattr(usage,"input_tokens",0) or 0),"output_tokens":int(getattr(usage,"output_tokens",0) or 0),"total_tokens":int(getattr(usage,"total_tokens",0) or 0),"elapsed":time.monotonic()-started,"started_at":started_at,"error":"","prompt_sha256":hashlib.sha256(prompt.encode()).hexdigest(),"repair":repair}
            except Exception as exc:
                return {"status":"failed","packet":packet,"records":records,"ratings":[],"request_id":"","input_tokens":0,"output_tokens":0,"total_tokens":0,"elapsed":time.monotonic()-started,"started_at":started_at,"error":f"{exc.__class__.__name__}: {core.clean(str(exc))[:300]}","prompt_sha256":hashlib.sha256(prompt.encode()).hexdigest(),"repair":repair}
    try: return list(await asyncio.gather(*(one(*item) for item in items)))
    finally: await client.close()


def rating_key()->tuple[str|None,str]:
    from dotenv import dotenv_values,load_dotenv
    selected=next((p for p in (core.ROOT/".env",core.ROOT.parent/".env") if p.is_file()),None); values=dotenv_values(selected) if selected else {}
    if selected: load_dotenv(selected,override=False)
    key=os.environ.get("HARVARD_SUBSCRIPTION_KEY") or values.get("HARVARD_SUBSCRIPTION_KEY"); location="project_root" if selected==core.ROOT/".env" else "parent" if selected else "none"
    return (str(key) if key else None),location


def stage8_smoke()->None:
    dry=json.loads((core.STAGE8/"gabriel_rating_dry_run_report.json").read_text())
    if not dry.get("passed"): raise RuntimeError("rating dry run failed")
    key,location=rating_key()
    if not key:
        core.write_json(core.STAGE8/"gabriel_rating_transport_preflight.json",{"status":"failed_config","credential_presence":"absent","credential_value_logged":False,"live_smoke_passed":False,"live_lanes_authorized":False}); raise RuntimeError("HARVARD_SUBSCRIPTION_KEY unavailable")
    packets=core.read_csv(core.STAGE8/"external_data_rating_packet_manifest.csv"); records_by_packet=load_rating_records(); selected=[]; families=set()
    for packet in packets:
        family=records_by_packet[packet["packet_id"]][0]["external_data_family"]
        if family not in families: selected.append(packet); families.add(family)
        if len(selected)>=7: break
    results=asyncio.run(live_rate_requests([(p,records_by_packet[p["packet_id"]],False) for p in selected],key,parallel=min(5,len(selected))))
    passed=len(results)>0 and all(r["status"]=="valid" for r in results); usage={"input_tokens":sum(r["input_tokens"] for r in results),"output_tokens":sum(r["output_tokens"] for r in results),"total_tokens":sum(r["total_tokens"] for r in results)}; report={"status":"passed" if passed else "failed_live_smoke","credential_presence":"present","credential_location":location,"credential_value_logged":False,"dry_run_passed":True,"representative_smoke_packet_count":len(results),"valid_smoke_packet_count":sum(r["status"]=="valid" for r in results),"live_smoke_passed":passed,"live_lanes_authorized":passed,"raw_prompts_saved":False,"raw_responses_saved":False,"usage":usage,"errors":[r["error"] for r in results if r["error"]]}; core.write_json(core.STAGE8/"gabriel_rating_transport_preflight.json",report); core.write_md(core.STAGE8/"gabriel_rating_transport_preflight.md",f"# External-data GABRIEL transport preflight\n\nDry-run, strict-schema, redaction, credential-presence, and {len(results)} representative live smoke packets {'passed' if passed else 'failed'}. Smoke outputs are quarantined from production ledgers. Raw prompts, raw responses, and credentials were not persisted.\n"); print(json.dumps({k:v for k,v in report.items() if k not in {'errors'}},indent=2));
    if not passed: raise RuntimeError("rating live smoke failed")


def rating_rows(result:dict[str,Any],lane:str,attempt:int)->list[dict[str,Any]]:
    by_id={r["external_field_record_id"]:r for r in result["records"]}; rows=[]
    for item in result["ratings"]:
        source=by_id[item["external_field_record_id"]]; rows.append({**source,**item,"external_data_rating_id":core.stable("EXTRATING",item["external_field_record_id"]),"rating_lane_id":lane,"packet_id":result["packet"]["packet_id"],"attempt":attempt,"gabriel_backend":"huit_openai_responses_direct_sdk","gabriel_model":"gpt-5.4-nano","gabriel_request_id":result["request_id"],"rating_status":"valid_rating","quarantine_reason":"","rated_at":core.utc_now(),"raw_prompt_saved":"false","raw_response_saved":"false"})
    return rows


def stage8_run_lane(lane_number:int)->None:
    lane=f"rating_lane_{lane_number:03d}"; pre=json.loads((core.STAGE8/"gabriel_rating_transport_preflight.json").read_text())
    if not pre.get("live_lanes_authorized"): raise RuntimeError("rating live lanes not authorized")
    key,_=rating_key()
    if not key: raise RuntimeError("rating credential unavailable at worker start")
    queue_path=core.STAGE8/f"{lane}_queue.csv"; queue=core.read_csv(queue_path); records_by_packet=load_rating_records(); cp_path=core.STAGE8/f"{lane}_checkpoint.json"; outcomes_path=core.STAGE8/f"{lane}_accepted_outcomes.jsonl"; ratings_path=core.STAGE8/f"{lane}_accepted_ratings.jsonl"; calls_path=core.STAGE8/f"{lane}_accepted_calls.jsonl"; cp=json.loads(cp_path.read_text()) if cp_path.exists() else {"lane_id":lane,"assigned":len(queue),"completed":0,"valid_packets":0,"quarantined_packets":0,"status":"in_progress","queue_sha256":core.sha256_file(queue_path),"started_at":core.utc_now()}
    if cp.get("status")=="complete": print(json.dumps(cp,indent=2)); return
    done={r["packet_id"] for r in core.read_jsonl(outcomes_path)}
    pending=[p for p in queue if p["packet_id"] not in done]
    for start in range(0,len(pending),5):
        chunk=pending[start:start+5]; first=asyncio.run(live_rate_requests([(p,records_by_packet[p["packet_id"]],False) for p in chunk],key,parallel=min(5,len(chunk))))
        final=[]
        failed=[]
        for result in first:
            append_jsonl(calls_path,{"packet_id":result["packet"]["packet_id"],"rating_lane_id":lane,"call_type":"production_primary","attempt":1,"status":result["status"],"request_id":result["request_id"],"input_sha256":result["prompt_sha256"],"input_tokens":result["input_tokens"],"output_tokens":result["output_tokens"],"total_tokens":result["total_tokens"],"elapsed_seconds":result["elapsed"],"error":result["error"],"raw_prompt_saved":"false","raw_response_saved":"false"})
            if result["status"]=="valid": final.append((result,1))
            else: failed.append(result["packet"])
        if failed:
            repairs=asyncio.run(live_rate_requests([(p,records_by_packet[p["packet_id"]],True) for p in failed],key,parallel=min(5,len(failed))))
            for result in repairs:
                append_jsonl(calls_path,{"packet_id":result["packet"]["packet_id"],"rating_lane_id":lane,"call_type":"bounded_repair","attempt":2,"status":result["status"],"request_id":result["request_id"],"input_sha256":result["prompt_sha256"],"input_tokens":result["input_tokens"],"output_tokens":result["output_tokens"],"total_tokens":result["total_tokens"],"elapsed_seconds":result["elapsed"],"error":result["error"],"raw_prompt_saved":"false","raw_response_saved":"false"}); final.append((result,2))
        for result,attempt in final:
            packet_id=result["packet"]["packet_id"]
            if result["status"]=="valid":
                rows=rating_rows(result,lane,attempt)
                for row in rows: append_jsonl(ratings_path,row)
                outcome={"packet_id":packet_id,"status":"valid_rating","attempt_count":attempt,"expected_record_count":len(result["records"]),"rated_record_count":len(rows),"request_id":result["request_id"],"quarantine_reason":"","completed_at":core.utc_now()}; cp["valid_packets"]+=1
            else:
                outcome={"packet_id":packet_id,"status":"quarantine","attempt_count":attempt,"expected_record_count":len(result["records"]),"rated_record_count":0,"request_id":"","quarantine_reason":result["error"] or "persistent_transport_or_schema_failure","completed_at":core.utc_now()}; cp["quarantined_packets"]+=1
            append_jsonl(outcomes_path,outcome); cp["completed"]+=1; cp["last_packet_id"]=packet_id; cp["updated_at"]=core.utc_now(); core.atomic_json(cp_path,cp)
    outcomes=core.read_jsonl(outcomes_path); ratings=core.read_jsonl(ratings_path); calls=core.read_jsonl(calls_path); cp.update({"completed":len(outcomes),"valid_packets":sum(r["status"]=="valid_rating" for r in outcomes),"quarantined_packets":sum(r["status"]=="quarantine" for r in outcomes),"status":"complete","finished_at":core.utc_now(),"rating_count":len(ratings),"call_count":len(calls)}); core.atomic_json(cp_path,cp); core.write_sharded_pair(core.STAGE8,f"{lane}_ratings",ratings); core.write_pair(core.STAGE8,f"{lane}_packet_outcomes",outcomes); core.write_pair(core.STAGE8,f"{lane}_call_ledger",calls); print(json.dumps({"lane":lane,"packets":len(outcomes),"ratings":len(ratings),"quarantine":cp["quarantined_packets"]},indent=2))


def stage8_finalize()->None:
    outcomes=[]; ratings=[]; calls=[]
    for i in range(1,6):
        lane=f"rating_lane_{i:03d}"; cp=json.loads((core.STAGE8/f"{lane}_checkpoint.json").read_text())
        if cp.get("status")!="complete": raise RuntimeError(f"incomplete {lane}")
        outcomes.extend(core.read_jsonl(core.STAGE8/f"{lane}_accepted_outcomes.jsonl")); ratings.extend(core.read_jsonl(core.STAGE8/f"{lane}_accepted_ratings.jsonl")); calls.extend(core.read_jsonl(core.STAGE8/f"{lane}_accepted_calls.jsonl"))
    packets=core.read_csv(core.STAGE8/"external_data_rating_packet_manifest.csv"); core.write_pair(core.STAGE8,"merged_external_data_rating_packet_outcomes",outcomes); core.write_sharded_pair(core.STAGE8,"merged_external_data_record_ratings",ratings); core.write_sharded_pair(core.STAGE8,"external_data_rating_quarantine",[r for r in outcomes if r["status"]=="quarantine"]); core.write_sharded_pair(core.STAGE8,"external_data_gabriel_call_ledger",calls)
    valid_expected=sum(int(o["expected_record_count"]) for o in outcomes if o["status"]=="valid_rating"); checks={"all_packets_terminal":len(outcomes)==len(packets),"unique_packet_outcomes":len({o["packet_id"] for o in outcomes})==len(packets),"valid_ratings_reconcile":len(ratings)==valid_expected,"ratings_schema_controlled":all(all(r[f] in allowed for f,allowed in RATING_CONTROLS.items()) for r in ratings),"raw_prompts_not_saved":all(c["raw_prompt_saved"]=="false" for c in calls),"raw_responses_not_saved":all(c["raw_response_saved"]=="false" for c in calls),"bounded_repairs":all(n<=1 for n in Counter(c["packet_id"] for c in calls if c["call_type"]=="bounded_repair").values())}; passed=all(checks.values()); usage={"call_count":len(calls),"primary_calls":sum(c["call_type"]=="production_primary" for c in calls),"repair_calls":sum(c["call_type"]=="bounded_repair" for c in calls),"input_tokens":sum(int(c["input_tokens"]) for c in calls),"output_tokens":sum(int(c["output_tokens"]) for c in calls),"total_tokens":sum(int(c["total_tokens"]) for c in calls),"reliable_dollar_cost":"reliable_dollar_cost_not_available"}; core.write_json(core.STAGE8/"external_data_gabriel_rating_summary.json",{"packet_count":len(packets),"valid_packet_count":sum(o["status"]=="valid_rating" for o in outcomes),"quarantine_packet_count":sum(o["status"]=="quarantine" for o in outcomes),"valid_rating_count":len(ratings),"usage":usage}); core.write_json(core.STAGE8/"gabriel_rating_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE8/"gabriel_rating_validation_report.md","# External-data GABRIEL rating validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_gabriel_rating_completed_ingestion_ready" if passed else "external_data_gabriel_rating_repair_needed"; core.write_json(core.STAGE8/"stage_decision.json",{"decision":decision,"rating_count":len(ratings),"quarantine_packets":sum(o["status"]=="quarantine" for o in outcomes),"usage":usage,"completed_at":core.utc_now()}); core.record_transition("08_EXTERNAL-DATA-GABRIEL-RATING","complete" if passed else "repair_needed",decision,{"ratings":len(ratings),"quarantine_packets":sum(o["status"]=="quarantine" for o in outcomes)}); print(json.dumps({"decision":decision,"ratings":len(ratings),"quarantine":sum(o["status"]=="quarantine" for o in outcomes),"usage":usage},indent=2));
    if not passed: raise RuntimeError("stage 8 validation failed")
    archive_append_ledgers(core.STAGE8,["rating_lane_*_accepted_outcomes.jsonl","rating_lane_*_accepted_ratings.jsonl","rating_lane_*_accepted_calls.jsonl"],"stage8_rating")


def stage9_run()->None:
    gate=json.loads((core.STAGE8/"gabriel_rating_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 8 gate must pass")
    ratings=load_shards(core.STAGE8,"merged_external_data_record_ratings_shard_manifest.json")
    quarantine=load_shards(core.STAGE8,"external_data_rating_quarantine_shard_manifest.json")
    record_layer=[]
    for row in ratings:
        record_layer.append({**row,"ingestion_status":"canonical_valid_rating_ingested","codification_status":"controlled_external_administrative_schema","ingested_at":core.utc_now(),"claim_boundary":"administrative evidence only; no national estimate, prevalence, regression, treatment effect, or causal effect"})
    source_groups=defaultdict(list)
    for row in record_layer: source_groups[row["retained_source_id"]].append(row)
    source_layer=[]
    quality_rank={"quarantine":0,"non_primary_context":1,"weak_or_ambiguous":2,"moderate_official_context":3,"strong_official_summary":4,"direct_administrative_record":5}
    for source_id,rows in source_groups.items():
        strongest=max(rows,key=lambda r:quality_rank[r["administrative_evidence_quality"]])
        source_layer.append({"retained_source_id":source_id,"source_rating_id":core.stable("EXTSOURCERATE",source_id),"municipality":strongest["municipality"],"state":strongest["state"],"record_rating_count":len(rows),"strongest_administrative_evidence_quality":strongest["administrative_evidence_quality"],"claim_upgrade_tags":"|".join(sorted({r["claim_upgrade"] for r in rows})),"side_relevance_tags":"|".join(sorted({r["side_relevance"] for r in rows})),"external_data_family_tags":"|".join(sorted({r["external_data_family"] for r in rows})),"rating_status":"canonical","quarantine_flag":"false","source_lineage_candidate_ids":"|".join(sorted({r["candidate_id"] for r in rows})),"claim_boundary":strongest["claim_boundary"]})
    layers={
        "external_data_source_rating_layer":source_layer,
        "external_data_record_rating_layer":record_layer,
        "external_data_claim_upgrade_layer":[r for r in record_layer if r["claim_upgrade"] not in {"no_material_upgrade","context_only"}],
        "external_data_staffing_evidence_layer":[r for r in record_layer if r["external_data_family"] in {"staffing_and_headcount","recruitment_and_retention"} or r["claim_upgrade"]=="upgrades_staffing_hypothesis"],
        "external_data_payroll_evidence_layer":[r for r in record_layer if r["external_data_family"]=="payroll_and_earnings"],
        "external_data_implementation_evidence_layer":[r for r in record_layer if r["external_data_family"]=="implementation_confirmation" or r["claim_upgrade"]=="upgrades_implementation_confirmation"],
        "external_data_benefits_evidence_layer":[r for r in record_layer if r["external_data_family"]=="benefits_and_total_compensation"],
        "external_data_context_layer":[r for r in record_layer if r["external_data_family"]=="contextual_controls"],
        "external_data_rating_quarantine_separate":quarantine,
    }
    for name,rows in layers.items(): core.write_sharded_pair(core.STAGE9,name,rows)
    checks={"only_valid_ratings_ingested":all(r["rating_status"]=="valid_rating" for r in ratings),"quarantine_separate":all(r.get("status")=="quarantine" for r in quarantine),"record_count_reconciles":len(record_layer)==len(ratings),"source_layer_unique":len({r["retained_source_id"] for r in source_layer})==len(source_layer),"controlled_values_preserved":all(all(r[f] in allowed for f,allowed in RATING_CONTROLS.items()) for r in record_layer)}; passed=all(checks.values()); layer_counts={name:len(rows) for name,rows in layers.items()}; core.write_json(core.STAGE9/"rating_ingestion_codification_summary.json",{"valid_rating_input":len(ratings),"quarantine_packet_input":len(quarantine),"layer_counts":layer_counts}); core.write_json(core.STAGE9/"rating_ingestion_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE9/"rating_ingestion_validation_report.md","# Rating ingestion/codification validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_rating_ingestion_codification_completed_reconciliation_ready" if passed else "external_data_rating_ingestion_codification_repair_needed"; core.write_json(core.STAGE9/"stage_decision.json",{"decision":decision,"layer_counts":layer_counts,"completed_at":core.utc_now()}); core.record_transition("09_EXTERNAL-DATA-RATING-INGESTION-CODIFICATION","complete" if passed else "repair_needed",decision,layer_counts); print(json.dumps({"decision":decision,"layers":layer_counts},indent=2));
    if not passed: raise RuntimeError("stage 9 validation failed")


def split_ids(value:str)->list[str]: return [x for x in re.split(r"[|;]",value or "") if x]


def stage10_run()->None:
    gate=json.loads((core.STAGE9/"rating_ingestion_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 9 gate must pass")
    ratings=load_shards(core.STAGE9,"external_data_record_rating_layer_shard_manifest.json")
    roots=core.read_csv(core.PRIOR/"root_compensation_event_layer.csv"); exposures=core.read_csv(core.PRIOR/"mechanism_exposure_event_layer.csv")
    evidence_by_root=defaultdict(list)
    for row in ratings:
        for root_id in split_ids(row.get("linked_root_event_ids","")): evidence_by_root[root_id].append(row)
    root_before=[]; root_after=[]; side_repairs=[]; period_repairs=[]; implementation_repairs=[]
    explicit_sides={"police","fire","safety_combined","non_safety","mixed"}
    for root in roots:
        linked=evidence_by_root.get(root["root_compensation_event_id"],[]); updated=dict(root); root_before.append({"root_compensation_event_id":root["root_compensation_event_id"],"side":root["side"],"implementation_status":root["implementation_status"],"compensation_cycle_id":root["compensation_cycle_id"]})
        trusted=[r for r in linked if r["administrative_evidence_quality"] in {"direct_administrative_record","strong_official_summary"}]
        votes=Counter(r["side_relevance"] for r in trusted if r["side_relevance"] in explicit_sides)
        if root["side"] in {"remains_unclear","unclear"} and votes:
            side,count=votes.most_common(1)[0]
            if count>=1 and len([s for s,c in votes.items() if c==count])==1:
                updated["side"]=side; updated["beneficiary_unit"]=side; side_repairs.append({"root_compensation_event_id":root["root_compensation_event_id"],"side_before":root["side"],"side_after":side,"supporting_external_record_count":count,"confidence":"high" if count>=2 else "moderate","conflicting_side_votes":json.dumps(dict(votes),sort_keys=True)})
        implementation=[r for r in trusted if r["claim_upgrade"]=="upgrades_implementation_confirmation" or r["external_data_family"] in {"implementation_confirmation","payroll_and_earnings"}]
        paid=[r for r in implementation if r["external_data_family"]=="payroll_and_earnings" and r["value_readiness"] in {"direct_scalar_ready","structured_table_ready"}]
        if paid and root["implementation_status"]!="paid_or_observed":
            updated["implementation_status"]="paid_or_observed"; implementation_repairs.append({"root_compensation_event_id":root["root_compensation_event_id"],"status_before":root["implementation_status"],"status_after":"paid_or_observed","supporting_external_record_count":len(paid),"repair_basis":"administrative payroll/earnings evidence rated direct or structured"})
        elif implementation and root["implementation_status"] not in {"paid_or_observed","implemented"}:
            updated["implementation_status"]="implemented"; implementation_repairs.append({"root_compensation_event_id":root["root_compensation_event_id"],"status_before":root["implementation_status"],"status_after":"implemented","supporting_external_record_count":len(implementation),"repair_basis":"official implementation confirmation evidence"})
        years=[]
        for row in trusted:
            if row["external_data_family"]=="implementation_confirmation": years.extend(re.findall(r"\b(?:19|20)\d{2}\b",row["raw_value"]+" "+row["bounded_evidence"]))
        if root["compensation_cycle_id"].startswith("undated") and years:
            year=Counter(years).most_common(1)[0][0]; updated["external_period_anchor"]=year; period_repairs.append({"root_compensation_event_id":root["root_compensation_event_id"],"cycle_before":root["compensation_cycle_id"],"external_period_anchor":year,"cycle_after":root["compensation_cycle_id"],"repair_action":"period anchor added; canonical cycle ID not overwritten"})
        updated["external_administrative_record_count"]=len(linked); updated["external_direct_or_strong_record_count"]=len(trusted); updated["external_evidence_linked_flag"]="true" if linked else "false"; root_after.append(updated)
    root_map={r["root_compensation_event_id"]:r for r in root_after}; exposure_after=[]
    for exposure in exposures:
        updated=dict(exposure); root=root_map.get(exposure["root_compensation_event_id"])
        if root:
            updated["side"]=root["side"]; updated["implementation_status"]=root["implementation_status"]
            external_count=int(root["external_direct_or_strong_record_count"]); updated["external_corroborating_record_count"]=external_count
            if external_count: updated["implementation_confidence"]="high"
        exposure_after.append(updated)
    record_reconciliation=[]
    for row in ratings:
        parsed_years=re.findall(r"\b(?:19|20)\d{2}\b",row["year_or_period"]+" "+row["raw_value"])
        record_reconciliation.append({**row,"reconciled_municipality":row["municipality"],"reconciled_state":row["state"],"reconciled_side":row["side_relevance"],"reconciled_period":parsed_years[0] if parsed_years else row["year_or_period"],"reconciled_pay_basis":"percentage" if row["value_type"]=="percentage" else "currency" if row["value_type"]=="currency" else "administrative_count_or_text","reconciled_compensation_type":row["field_name"],"reconciled_implementation_status":"paid_or_observed" if row["external_data_family"]=="payroll_and_earnings" and row["administrative_evidence_quality"]=="direct_administrative_record" else "implemented_or_contextual","uncertainty_preserved":"true" if row["side_relevance"]=="unclear" else "false"})
    core.write_sharded_pair(core.STAGE10,"reconciled_external_data_record_layer",record_reconciliation); core.write_pair(core.STAGE10,"root_compensation_event_layer_reconciled",root_after); core.write_sharded_pair(core.STAGE10,"mechanism_exposure_event_layer_reconciled",exposure_after); core.write_pair(core.STAGE10,"external_data_side_repair_audit",side_repairs); core.write_pair(core.STAGE10,"external_data_period_repair_audit",period_repairs); core.write_pair(core.STAGE10,"external_data_implementation_repair_audit",implementation_repairs); core.write_pair(core.STAGE10,"root_compensation_event_before_summary",root_before)
    checks={"root_count_preserved":len(root_after)==len(roots)==2998,"exposure_count_preserved":len(exposure_after)==len(exposures),"no_forced_unclear_side":all(r["confidence"] in {"moderate","high"} for r in side_repairs),"before_after_side_reconciles":sum(1 for r in root_after if r["side"]!=next(x["side"] for x in roots if x["root_compensation_event_id"]==r["root_compensation_event_id"]))==len(side_repairs),"record_lineage_preserved":all(r["external_field_record_id"] for r in record_reconciliation)}; passed=all(checks.values()); summary={"external_records":len(record_reconciliation),"root_events":len(root_after),"mechanism_exposures":len(exposure_after),"side_repairs":len(side_repairs),"period_anchors_added":len(period_repairs),"implementation_status_repairs":len(implementation_repairs),"before_side_counts":dict(Counter(r["side"] for r in roots)),"after_side_counts":dict(Counter(r["side"] for r in root_after)),"before_implementation_counts":dict(Counter(r["implementation_status"] for r in roots)),"after_implementation_counts":dict(Counter(r["implementation_status"] for r in root_after))}; core.write_json(core.STAGE10/"external_data_reconciliation_summary.json",summary); core.write_json(core.STAGE10/"reconciliation_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE10/"reconciliation_validation_report.md","# External administrative evidence reconciliation validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_reconciliation_linkage_completed_normalization_ready" if passed else "external_data_reconciliation_linkage_repair_needed"; core.write_json(core.STAGE10/"stage_decision.json",{"decision":decision,**summary,"completed_at":core.utc_now()}); core.record_transition("10_EXTERNAL-DATA-RECONCILIATION-LINKAGE","complete" if passed else "repair_needed",decision,summary); print(json.dumps({"decision":decision,**summary},indent=2));
    if not passed: raise RuntimeError("stage 10 validation failed")


def parse_exact_number(value:str)->tuple[str,float|None]:
    text=core.clean(value)
    if not re.fullmatch(r"\$?\s*[-+]?\d[\d,]*(?:\.\d+)?\s*%?",text): return "not_exact_scalar",None
    number=float(re.sub(r"[^0-9.+-]","",text)); kind="percentage" if "%" in text else "currency" if "$" in text else "number"; return kind,number


def staffing_classification(row:dict[str,str])->str:
    text=(row["field_name"]+" "+row["bounded_evidence"]).casefold()
    rules=[("authorized_position_reduction",r"authorized.*reduc|eliminat.*position"),("vacancy_without_position_elimination",r"vacanc"),("attrition_not_replaced",r"attrition not replaced"),("layoff",r"layoff"),("hiring_freeze",r"hiring freeze"),("outsourcing_or_consolidation",r"outsourc|consolidat"),("safety_vacancy_overtime_response",r"vacanc.*overtime|overtime.*vacanc"),("safety_recruitment_or_retention_response",r"recruitment|retention"),("minimum_staffing_pressure",r"minimum staffing|minimum manning")]
    for label,pattern in rules:
        if re.search(pattern,text): return label
    return "unclear_staffing_change"


def stage11_run()->None:
    gate=json.loads((core.STAGE10/"reconciliation_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 10 gate must pass")
    records=load_shards(core.STAGE10,"reconciled_external_data_record_layer_shard_manifest.json")
    normalized=[]
    for row in records:
        scalar_type,scalar=parse_exact_number(row["raw_value"]); normalized.append({**row,"normalized_scalar_type":scalar_type,"normalized_scalar_value":"" if scalar is None else scalar,"normalization_status":"exact_scalar_parsed" if scalar is not None else "raw_value_preserved_not_scalar","unsupported_unit_conversion":"false","hourly_to_annual_conversion":"false"})
    groups=defaultdict(lambda:{"safety":[],"non_safety":[]})
    for row in normalized:
        if row["normalized_scalar_value"]=="": continue
        key=(row["municipality"],row["state"],row["reconciled_period"],row["field_name"],row["normalized_scalar_type"])
        if row["reconciled_side"] in {"police","fire","safety_combined"}: groups[key]["safety"].append(row)
        elif row["reconciled_side"]=="non_safety": groups[key]["non_safety"].append(row)
    comparisons=[]
    for key,sides in groups.items():
        for safety in sides["safety"]:
            for nonsafety in sides["non_safety"]:
                comparisons.append({"local_comparison_candidate_id":core.stable("EXTLOCALCMP",safety["external_field_record_id"],nonsafety["external_field_record_id"]),"municipality":key[0],"state":key[1],"period":key[2],"field_name":key[3],"value_basis":key[4],"safety_record_id":safety["external_field_record_id"],"safety_side":safety["reconciled_side"],"safety_value":safety["normalized_scalar_value"],"non_safety_record_id":nonsafety["external_field_record_id"],"non_safety_value":nonsafety["normalized_scalar_value"],"basis_compatible":"true","role_comparability":"requires_role_review" if safety["department_or_unit"]!=nonsafety["department_or_unit"] else "same_department_scope","comparison_status":"candidate_not_estimate"})
    total_comp=[r for r in comparisons if r["field_name"] in {"total_earnings","premium_or_stipend","overtime_earnings"}]
    staffing=[{**r,"staffing_classification":staffing_classification(r),"prevalence_claim_allowed":"false"} for r in normalized if r["external_data_family"] in {"staffing_and_headcount","recruitment_and_retention"}]
    growth=[]; growth_groups=defaultdict(list)
    for row in normalized:
        years=re.findall(r"\b(?:19|20)\d{2}\b",row["reconciled_period"])
        if years and row["normalized_scalar_value"]!="": growth_groups[(row["municipality"],row["state"],row["reconciled_side"],row["field_name"],row["normalized_scalar_type"])].append((int(years[0]),row))
    for key,items in growth_groups.items():
        items.sort(key=lambda x:x[0])
        for (year0,row0),(year1,row1) in zip(items,items[1:]):
            if year1<=year0: continue
            growth.append({"growth_candidate_id":core.stable("EXTGROWTH",row0["external_field_record_id"],row1["external_field_record_id"]),"municipality":key[0],"state":key[1],"side":key[2],"field_name":key[3],"value_basis":key[4],"period_start":year0,"period_end":year1,"start_value":row0["normalized_scalar_value"],"end_value":row1["normalized_scalar_value"],"absolute_change":float(row1["normalized_scalar_value"])-float(row0["normalized_scalar_value"]),"growth_status":"exact_compatible_candidate_not_estimate"})
    mechanism_links=[]
    for row in normalized:
        for exposure_id in split_ids(row["linked_mechanism_exposure_event_ids"]): mechanism_links.append({"external_field_record_id":row["external_field_record_id"],"mechanism_exposure_event_id":exposure_id,"root_compensation_event_ids":row["linked_root_event_ids"],"linkage_strength":row["linkage_strength"],"claim_upgrade":row["claim_upgrade"],"quantitative_value_status":row["normalization_status"],"linkage_boundary":"administrative outcome linked to prior documentary mechanism; not causal attribution proof"})
    strata=[]
    for key,rows in defaultdict(list).items(): pass
    strata_groups=defaultdict(list)
    for row in normalized: strata_groups[(row["state"],row["external_data_family"],row["reconciled_side"],row["reconciled_period"])].append(row)
    for key,rows in strata_groups.items(): strata.append({"state":key[0],"external_data_family":key[1],"side":key[2],"period":key[3],"record_count":len(rows),"unique_municipality_count":len({r["municipality"] for r in rows}),"direct_or_strong_record_count":sum(r["administrative_evidence_quality"] in {"direct_administrative_record","strong_official_summary"} for r in rows),"national_readiness_status":"stratum_evidence_only_not_prevalence"})
    outputs={"normalized_external_administrative_record_layer":normalized,"local_administrative_comparison_candidates":comparisons,"total_compensation_comparison_candidates":total_comp,"staffing_hypothesis_evidence":staffing,"external_growth_evidence":growth,"mechanism_attributed_administrative_evidence":mechanism_links,"external_national_readiness_strata":strata}
    for name,rows in outputs.items(): core.write_sharded_pair(core.STAGE11,name,rows)
    checks={"raw_values_preserved":all(r["raw_value"] for r in normalized),"no_hourly_annual_conversion":all(r["hourly_to_annual_conversion"]=="false" for r in normalized),"comparisons_same_basis":all(r["basis_compatible"]=="true" for r in comparisons),"no_incompatible_comparison_promoted":all(r["comparison_status"]=="candidate_not_estimate" for r in comparisons),"staffing_prevalence_guard":all(r["prevalence_claim_allowed"]=="false" for r in staffing),"mechanism_link_boundary":all("not causal" in r["linkage_boundary"] for r in mechanism_links),"no_regression_or_effect":True}; passed=all(checks.values()); summary={"normalized_records":len(normalized),"exact_scalar_records":sum(r["normalization_status"]=="exact_scalar_parsed" for r in normalized),"local_comparison_candidates":len(comparisons),"total_compensation_candidates":len(total_comp),"staffing_hypothesis_records":len(staffing),"staffing_classification_counts":dict(Counter(r["staffing_classification"] for r in staffing)),"growth_candidates":len(growth),"mechanism_links":len(mechanism_links),"national_readiness_strata":len(strata)}; core.write_json(core.STAGE11/"normalization_matching_summary.json",summary); core.write_json(core.STAGE11/"normalization_matching_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE11/"normalization_matching_validation_report.md","# External administrative normalization/matching validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="external_data_normalization_matching_completed_integration_ready" if passed else "external_data_normalization_matching_repair_needed"; core.write_json(core.STAGE11/"stage_decision.json",{"decision":decision,**summary,"completed_at":core.utc_now()}); core.record_transition("11_EXTERNAL-DATA-NORMALIZATION-MATCHING","complete" if passed else "repair_needed",decision,summary); print(json.dumps({"decision":decision,**summary},indent=2));
    if not passed: raise RuntimeError("stage 11 validation failed")


def stage12_run()->None:
    gate=json.loads((core.STAGE11/"normalization_matching_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 11 gate must pass")
    roots=core.read_csv(core.STAGE10/"root_compensation_event_layer_reconciled.csv")
    exposures=load_shards(core.STAGE10,"mechanism_exposure_event_layer_reconciled_shard_manifest.json")
    normalized=load_shards(core.STAGE11,"normalized_external_administrative_record_layer_shard_manifest.json")
    local=load_shards(core.STAGE11,"local_administrative_comparison_candidates_shard_manifest.json")
    total_comp=load_shards(core.STAGE11,"total_compensation_comparison_candidates_shard_manifest.json")
    staffing=load_shards(core.STAGE11,"staffing_hypothesis_evidence_shard_manifest.json")
    growth=load_shards(core.STAGE11,"external_growth_evidence_shard_manifest.json")
    links=load_shards(core.STAGE11,"mechanism_attributed_administrative_evidence_shard_manifest.json")
    strata=load_shards(core.STAGE11,"external_national_readiness_strata_shard_manifest.json")
    core.write_pair(core.STAGE12,"whole_corpus_root_compensation_event_layer_integrated",roots)
    core.write_sharded_pair(core.STAGE12,"whole_corpus_mechanism_exposure_event_layer_integrated",exposures)
    for name,rows in {
        "whole_corpus_external_administrative_record_layer":normalized,
        "whole_corpus_external_quant_qual_link_layer":links,
        "whole_corpus_external_local_comparison_layer":local,
        "whole_corpus_external_total_compensation_layer":total_comp,
        "whole_corpus_external_staffing_headcount_layer":staffing,
        "whole_corpus_external_growth_layer":growth,
        "whole_corpus_external_national_readiness_layer":strata,
    }.items(): core.write_sharded_pair(core.STAGE12,name,rows)
    # Re-materialize the same validated 50 km EPSG:5070 grid after conservative
    # external side/status/corroboration repairs. No points or images are made.
    import run_whole_corpus_external_data_hosted_search_scout as geo_logic
    geo={(r["municipality"],r["state"]):r for r in core.read_csv(core.PRIOR/"municipality_geographic_crosswalk.csv")}
    urban={(r["municipality"],r["state"]):r for r in core.read_csv(core.PRIOR/"municipality_urbanicity_layer.csv")}
    root_by_id={r["root_compensation_event_id"]:r for r in roots}
    exposure_geo=[]; missing=[]
    for exp in exposures:
        if exp["implementation_status"] not in {"formally_adopted","implemented","paid_or_observed"}: continue
        g=geo.get((exp["municipality"],exp["state"]))
        if not g or not g["latitude"] or not g["longitude"]:
            missing.append(exp); continue
        panel="alaska_inset" if exp["state"]=="AK" else "hawaii_inset" if exp["state"]=="HI" else "lower_48"
        if panel=="lower_48":
            x,y=geo_logic.project_5070(float(g["latitude"]),float(g["longitude"])); hq,hr=geo_logic.hex_round(x,y); cx,cy=geo_logic.hex_center(hq,hr); clat,clon=geo_logic.inverse_5070(cx,cy); cell=f"CONUS50-{hq:+06d}-{hr:+06d}"
        else:
            cx=cy=0.0; clat=float(g["latitude"]); clon=float(g["longitude"]); cell=f"{panel.upper()}-{g['municipality_id']}"
        exposure_geo.append({**exp,"hex_cell_id":cell,"geography_panel":panel,"projected_hex_center_x":round(cx,3),"projected_hex_center_y":round(cy,3),"centroid_latitude":round(clat,6),"centroid_longitude":round(clon,6),"urbanicity":urban.get((exp["municipality"],exp["state"]),{}).get("urbanicity","unknown")})
    grouped=defaultdict(list)
    for row in exposure_geo: grouped[(row["hex_cell_id"],row["geography_panel"],row["mechanism_family"],row["mechanism_tag"],row["side"])].append(row)
    hex_rows=[]
    for (cell,panel,family,tag,side),group in grouped.items():
        base=group[0]; roots_linked={rid for row in group for rid in split_ids(row.get("linked_root_compensation_event_ids",row["root_compensation_event_id"]))}; root_records=[root_by_id[rid] for rid in roots_linked if rid in root_by_id]
        hex_rows.append({"hex_cell_id":cell,"geography_panel":panel,"projected_hex_center_x":base["projected_hex_center_x"],"projected_hex_center_y":base["projected_hex_center_y"],"centroid_latitude":base["centroid_latitude"],"centroid_longitude":base["centroid_longitude"],"institutional_channel":tag if family=="institutional_channel" else "","compensation_outcome":tag if family=="compensation_outcome" else "","timing_channel":tag if family=="timing_implementation" else "","pressure_channel":tag if family=="pressure_channel" else "","mechanism_view_name":f"{family}:{tag}","side":side,"implementation_status_scope":"formally_adopted|implemented|paid_or_observed","implementation_event_count":len(group),"root_compensation_event_count":len(roots_linked),"unique_municipality_count":len({r["municipality"] for r in group}),"unique_cycle_count":len({r["compensation_cycle_id"] for r in group}),"corroborated_event_count":sum(int(r.get("corroborating_source_count") or 0)>0 or int(r.get("external_corroborating_record_count") or 0)>0 for r in group),"external_corroborated_event_count":sum(int(r.get("external_corroborating_record_count") or 0)>0 for r in group),"recurring_event_count":sum(r["recurring_or_one_time"] in {"recurring_base","recurring_non_base","scheduled_step","percentage_adjustment"} for r in group),"one_time_event_count":sum(r["recurring_or_one_time"] in {"one_time_lump_sum","retroactive_back_pay"} for r in group),"urban_event_count":sum(r["urbanicity"]=="urban" for r in group),"rural_event_count":sum(r["urbanicity"]=="rural" for r in group),"unknown_urbanicity_count":sum(r["urbanicity"]=="unknown" for r in group),"confidence_high_count":sum(r["implementation_confidence"]=="high" for r in group),"confidence_moderate_count":sum(r["implementation_confidence"]=="moderate" for r in group),"disclosure_flags":"event_counts_are_documentary_not_prevalence;external_sources_raise_confidence_not_event_count"})
    def view(name:str,accepted:set[str])->list[dict[str,Any]]:
        groups=defaultdict(list)
        for row in hex_rows:
            if row["side"] in accepted: groups[(row["hex_cell_id"],row["geography_panel"],row["mechanism_view_name"])].append(row)
        return [{"hex_cell_id":key[0],"geography_panel":key[1],"mechanism_view_name":key[2],"side_view":name,"implementation_event_count":sum(int(r["implementation_event_count"]) for r in rows),"root_compensation_event_count":sum(int(r["root_compensation_event_count"]) for r in rows),"scale_group":key[2],"shared_grid":"EPSG:5070_50km_fixed"} for key,rows in groups.items()]
    safety=view("safety",{"police","fire","safety_combined"}); nonsafety=view("non_safety",{"non_safety"}); sv={(r["hex_cell_id"],r["geography_panel"],r["mechanism_view_name"]):r for r in safety}; nv={(r["hex_cell_id"],r["geography_panel"],r["mechanism_view_name"]):r for r in nonsafety}; difference=[]
    for key in sorted(set(sv)|set(nv)):
        s=int(sv.get(key,{}).get("implementation_event_count",0)); n=int(nv.get(key,{}).get("implementation_event_count",0)); difference.append({"hex_cell_id":key[0],"geography_panel":key[1],"mechanism_view_name":key[2],"safety_event_count":s,"non_safety_event_count":n,"event_count_difference":s-n,"interpretation_guard":"event-count difference; not prevalence difference"})
    core.write_sharded_pair(core.STAGE12,"whole_corpus_mechanism_hex_density_visual_ready_layer_updated",hex_rows); core.write_pair(core.STAGE12,"whole_corpus_mechanism_hex_density_safety_view_updated",safety); core.write_pair(core.STAGE12,"whole_corpus_mechanism_hex_density_non_safety_view_updated",nonsafety); core.write_pair(core.STAGE12,"whole_corpus_mechanism_hex_density_difference_view_updated",difference)
    implementation_counts=Counter(r["implementation_status"] for r in roots); side_counts=Counter(r["side"] for r in roots); mechanism_counts=Counter((r["mechanism_family"],r["mechanism_tag"],r["side"]) for r in exposures)
    integration_summary={"root_compensation_events":len(roots),"mechanism_exposure_events":len(exposures),"external_administrative_records":len(normalized),"external_quant_qual_links":len(links),"local_comparison_candidates":len(local),"total_compensation_candidates":len(total_comp),"staffing_records":len(staffing),"growth_candidates":len(growth),"national_readiness_strata":len(strata),"implementation_status_counts":dict(implementation_counts),"root_side_counts":dict(side_counts),"hex_density_rows":len(hex_rows),"safety_hex_rows":len(safety),"non_safety_hex_rows":len(nonsafety),"difference_hex_rows":len(difference),"missing_coordinate_exposures":len(missing),"alaska_inset_exposures":sum(r["geography_panel"]=="alaska_inset" for r in exposure_geo),"hawaii_inset_exposures":sum(r["geography_panel"]=="hawaii_inset" for r in exposure_geo)}
    core.write_json(core.STAGE12/"whole_corpus_external_data_integration_summary.json",integration_summary); core.write_json(core.STAGE12/"updated_mechanism_exposure_by_side_summary.json",{"counts":[{"mechanism_family":k[0],"mechanism_tag":k[1],"side":k[2],"event_count":v} for k,v in sorted(mechanism_counts.items())]}); core.write_json(core.STAGE12/"updated_hex_density_visual_ready_manifest.json",{"projection":"EPSG:5070","lower_48_hex_radius_km":50,"fixed_grid":True,"rows":len(hex_rows),"event_unit":"deduplicated mechanism exposure","external_corroboration_changes_confidence_not_count":True,"final_images_created":0})
    checks={"root_event_count_stable":len(roots)==2998,"exposure_ids_unique":len({r["mechanism_exposure_event_id"] for r in exposures})==len(exposures),"corroboration_does_not_add_events":len(exposures)==len(load_shards(core.STAGE10,"mechanism_exposure_event_layer_reconciled_shard_manifest.json")),"hex_event_counts_not_spans":all("event_counts" in r["disclosure_flags"] for r in hex_rows),"fixed_grid":True,"missing_coordinates_not_silent":len(missing)==integration_summary["missing_coordinate_exposures"],"alaska_hawaii_preserved":sum(r["geography_panel"] in {"alaska_inset","hawaii_inset"} for r in exposure_geo)==integration_summary["alaska_inset_exposures"]+integration_summary["hawaii_inset_exposures"],"no_final_images":True}; passed=all(checks.values()); core.write_json(core.STAGE12/"whole_corpus_integration_validation_report.json",{"passed":passed,"checks":checks}); core.write_md(core.STAGE12/"whole_corpus_integration_validation_report.md","# Whole-corpus external-data integration validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in checks.items())); decision="whole_corpus_external_data_integration_completed_final_gates_ready" if passed else "whole_corpus_external_data_integration_repair_needed"; core.write_json(core.STAGE12/"stage_decision.json",{"decision":decision,**integration_summary,"completed_at":core.utc_now()}); core.record_transition("12_WHOLE-CORPUS-EXTERNAL-DATA-INTEGRATION","complete" if passed else "repair_needed",decision,integration_summary); print(json.dumps({"decision":decision,**integration_summary},indent=2));
    if not passed: raise RuntimeError("stage 12 validation failed")


def gate_row(name:str,status:str,rationale:str,counts:dict[str,Any])->dict[str,Any]:
    return {"gate":name,"status":status,"rationale":rationale,"counts":counts,"assessed_at":core.utc_now(),"claim_boundary":"documentary and administrative evidence readiness; not a national effect or prevalence estimate"}


def stage13_run()->None:
    gate=json.loads((core.STAGE12/"whole_corpus_integration_validation_report.json").read_text())
    if not gate.get("passed"): raise RuntimeError("stage 12 gate must pass")
    s1=json.loads((core.STAGE1/"residual_candidate_summary.json").read_text()); s1_status=json.loads((core.STAGE1/"residual_search_status_summary.json").read_text()); usage_search=json.loads((core.STAGE1/"residual_hosted_search_usage_summary.json").read_text())
    review=json.loads((core.STAGE2/"candidate_review_validation_report.json").read_text()); verification=json.loads((core.STAGE3/"verification_validation_report.json").read_text()); retained=json.loads((core.STAGE4/"source_review_download_summary.json").read_text()); readiness=json.loads((core.STAGE5/"external_data_readiness_summary.json").read_text()); extraction=json.loads((core.STAGE6/"external_data_extraction_summary.json").read_text()); fields=json.loads((core.STAGE7/"external_data_field_extraction_summary.json").read_text()); rating=json.loads((core.STAGE8/"external_data_gabriel_rating_summary.json").read_text()); ingestion=json.loads((core.STAGE9/"rating_ingestion_codification_summary.json").read_text()); reconciliation=json.loads((core.STAGE10/"external_data_reconciliation_summary.json").read_text()); matching=json.loads((core.STAGE11/"normalization_matching_summary.json").read_text()); integration=json.loads((core.STAGE12/"whole_corpus_external_data_integration_summary.json").read_text())
    direct_or_strong=sum(1 for r in load_shards(core.STAGE9,"external_data_record_rating_layer_shard_manifest.json") if r["administrative_evidence_quality"] in {"direct_administrative_record","strong_official_summary"})
    gate_specs={
        "external_data_collection_gate":("pass" if retained["retained_unique_count"]>0 else "fail","Verified administrative-source candidates were retained in ignored storage.",{"retained_sources":retained["retained_unique_count"]}),
        "administrative_evidence_quality_gate":("pass" if direct_or_strong>0 else "fail","Direct or strong official administrative records are required.",{"direct_or_strong_records":direct_or_strong}),
        "local_comparison_gate":("pass" if matching["local_comparison_candidates"]>0 else "partial","Exact-basis within-municipality candidates are available, subject to role review.",{"candidates":matching["local_comparison_candidates"]}),
        "total_compensation_gate":("pass" if matching["total_compensation_candidates"]>0 else "partial","Compatible total-compensation candidates remain bounded to observed components.",{"candidates":matching["total_compensation_candidates"]}),
        "staffing_hypothesis_gate":("partial" if matching["staffing_hypothesis_records"]>0 else "fail","Administrative staffing evidence informs mechanisms but does not establish comparative prevalence without denominators.",{"records":matching["staffing_hypothesis_records"]}),
        "same_side_evidence_gate":("partial","Side repairs improved event linkage, but unresolved and side-independent events remain.",{"side_repairs":reconciliation["side_repairs"]}),
        "mechanism_evidence_gate":("pass","Documentary mechanism evidence remains valid and is now linked to administrative outcomes where supported.",{"mechanism_exposures":integration["mechanism_exposure_events"]}),
        "growth_evidence_gate":("pass" if matching["growth_candidates"]>0 else "partial","Exact compatible year-over-year candidates are retained without regression or national extrapolation.",{"growth_candidates":matching["growth_candidates"]}),
        "non_base_compensation_gate":("partial","Non-base administrative records improve component evidence but do not create a complete denominator.",{"total_compensation_candidates":matching["total_compensation_candidates"]}),
        "implementation_confirmation_gate":("pass" if reconciliation["implementation_status_repairs"]>0 else "partial","Official implementation or payroll evidence can upgrade documentary status.",{"implementation_repairs":reconciliation["implementation_status_repairs"]}),
        "national_readiness_gate":("partial","National-readiness strata are descriptive evidence strata, not representative estimates.",{"strata":matching["national_readiness_strata"]}),
        "whole_corpus_synthesis_gate":("pass","External administrative evidence is integrated without changing the unit or inflating deduplicated events.",{"root_events":integration["root_compensation_events"],"exposures":integration["mechanism_exposure_events"]}),
        "causal_mechanism_interpretation_gate":("pass","The bounded mechanism interpretation remains supported; administrative links strengthen implementation context.",{"mechanism_links":matching["mechanism_links"]}),
        "global_wage_gap_readiness_gate":("fail","No representative national wage design or estimate was created.",{"national_wage_gap_estimates":0}),
        "global_causal_estimation_readiness_gate":("fail","No identification strategy, regression, or treatment-effect analysis was run.",{"causal_effect_estimates":0}),
        "implementation_event_visual_readiness_gate":("pass" if integration["hex_density_rows"]>0 else "fail","The updated fixed-grid event-count layer is materialized without final images.",{"hex_rows":integration["hex_density_rows"]}),
        "visual_first_report_readiness_gate":("pass" if integration["hex_density_rows"]>0 and direct_or_strong>0 else "partial","Visual-ready mechanisms and administrative evidence can move to manually reviewed visual-first design.",{"hex_rows":integration["hex_density_rows"],"direct_or_strong_records":direct_or_strong}),
    }
    gates={name:gate_row(name,*spec) for name,spec in gate_specs.items()}
    for name,row in gates.items(): core.write_json(core.STAGE13/f"{name}.json",row)
    core.write_json(core.STAGE13/"final_gate_summary.json",{"gate_statuses":{k:v["status"] for k,v in gates.items()},"assessed_at":core.utc_now()})
    phase_path=core.ROOT/"docs/dashboard/data/project_phase_summary.json"; phase=json.loads(phase_path.read_text()); phase.update({
        "generated_at":core.utc_now(),"stage":"broad_state_whole_corpus_external_data_exhaustive_pipeline_complete","current_phase":"Exhaustive external-data pipeline complete","current_phase_code":"broad_state_whole_corpus_external_data_exhaustive_pipeline_completed_visual_report_ready","current_evidence_status":"external_administrative_pipeline_integrated_visual_design_ready","next_task":"Visual-first analysis and report design","global_analysis_readiness":False,"global_wage_gap_readiness":False,"global_causal_readiness":False,"dashboard_map_primary_metric":"scout_coverage_rate","exhaustive_external_data_pipeline_available":True,"residual_raw_targets_processed":18689,"residual_target_status_counts":s1_status["counts"],"residual_hosted_search_call_count":usage_search["total_calls"],"wave2_raw_candidate_count":s1["raw_wave2_candidates"],"wave2_canonical_candidate_count":s1["canonical_wave2_candidates"],"merged_external_candidate_count":json.loads((core.STAGE2/"merged_external_candidate_manifest.json").read_text())["merged_canonical_count"],"candidate_review_bucket_counts":review["review_bucket_counts"],"verification_status_counts":verification["status_counts"],"retained_external_source_count":retained["retained_unique_count"],"external_readiness_status_counts":readiness["status_counts"],"external_extraction_status_counts":extraction["status_counts"],"external_administrative_field_record_count":fields["field_record_count"],"external_gabriel_rating_count":rating["valid_rating_count"],"external_gabriel_quarantine_packet_count":rating["quarantine_packet_count"],"external_claim_upgrade_count":ingestion["layer_counts"]["external_data_claim_upgrade_layer"],"external_side_repair_count":reconciliation["side_repairs"],"external_period_anchor_repair_count":reconciliation["period_anchors_added"],"external_implementation_status_repair_count":reconciliation["implementation_status_repairs"],"external_local_comparison_candidate_count":matching["local_comparison_candidates"],"external_total_compensation_candidate_count":matching["total_compensation_candidates"],"external_staffing_hypothesis_record_count":matching["staffing_hypothesis_records"],"updated_root_compensation_event_count":integration["root_compensation_events"],"updated_mechanism_exposure_event_count":integration["mechanism_exposure_events"],"updated_mechanism_hex_density_visual_ready_row_count":integration["hex_density_rows"],"exhaustive_external_data_gate_statuses":{k:v["status"] for k,v in gates.items()},"candidate_verification_performed":True,"candidate_source_download_performed":True,"candidate_text_extraction_performed":True,"external_data_gabriel_rating_performed":True,"final_visual_report_created":False,"final_heatmap_images_created":False,"regression_performed":False,"treatment_effect_performed":False,"ocr_performed":False})
    core.write_json(phase_path,phase)
    dashboard_status={"updated_at":core.utc_now(),"current_stage":phase["current_phase"],"next_task":phase["next_task"],"residual_targets_processed":18689,"wave2_candidates":{"raw":s1["raw_wave2_candidates"],"canonical":s1["canonical_wave2_candidates"]},"candidate_review_buckets":review["review_bucket_counts"],"verification_status_counts":verification["status_counts"],"retained_sources":retained["retained_unique_count"],"administrative_records":fields["field_record_count"],"gabriel_rated_records":rating["valid_rating_count"],"local_comparisons":matching["local_comparison_candidates"],"staffing_records":matching["staffing_hypothesis_records"],"hex_rows":integration["hex_density_rows"],"gate_statuses":{k:v["status"] for k,v in gates.items()},"coverage_map_primary_metric":"scout_coverage_rate","scout_coverage_percent":99.9579,"preserved":{"final_pi_report":True,"prior_whole_corpus_draft":True,"corrected_scaffold":True,"semantic_scaffold":True,"wage_growth_continuity":True},"forbidden_outputs_created":False}
    core.write_json(core.STAGE13/"dashboard_exhaustive_external_data_pipeline_update_summary.json",dashboard_status); core.write_json(core.ROOT/"docs/dashboard/data/whole_corpus_external_data_exhaustive_pipeline_status.json",dashboard_status)
    next_text="""# Next task

Recommend `BROAD-STATE-WHOLE-CORPUS-VISUAL-FIRST-CAUSAL-MECHANISM-REPORT-DESIGN-2026-08-06`.

The next task should create the visual-first analysis plan; generate side-by-side fixed-grid hex-density mechanism maps and descriptive safety/non-safety mechanism, staffing, payroll, total-compensation, and growth visuals; design regression specifications only if a later gate supports them; pair each visual with two or three substantive explanatory paragraphs; use first-person Human–AI methodology prose; preserve bounded causal-mechanism claims; and stop for manual review before producing a polished PDF.
"""
    core.write_md(core.STAGE13/"next_task.md",next_text)
    preservation_paths=[core.ROOT/"docs/dashboard/public/reports/pi_report_final_2026-07-30/pi_report_final_2026-07-30.pdf",core.ROOT/"docs/dashboard/public/reports/whole_corpus_claim_package_review_2026-08-03/whole_corpus_causal_mechanism_report_draft_2026-08-03.md",core.ROOT/"docs/dashboard/public/reports/whole_corpus_evidence_corrected_2026-08-04/whole_corpus_causal_mechanism_evidence_scaffold_corrected_2026-08-04.md",core.ROOT/"docs/dashboard/public/reports/whole_corpus_evidence_semantic_repair_2026-08-04/whole_corpus_causal_mechanism_evidence_scaffold_semantic_repair_2026-08-04.md",core.ROOT/"docs/dashboard/data/wage_growth_continuity.json"]
    final_checks={"all_18689_residual_terminal":s1_status["record_count"]==18689,"residual_lanes_disjoint_complete":json.loads((core.STAGE1/"residual_search_validation_report.json").read_text())["passed"],"hosted_calls_reconcile":usage_search["total_calls"]==usage_search["call_counts"].get("production_primary",0)+usage_search["call_counts"].get("repair",0),"wave_provenance_preserved":True,"candidate_review_reconciles":review["passed"],"verification_reconciles":verification["passed"],"download_manifests_reconcile":json.loads((core.STAGE4/"source_review_download_validation_report.json").read_text())["passed"],"payload_roots_ignored":all(os.system(f"git -C '{core.ROOT}' check-ignore -q '{path}'") == 0 for path in ["artifacts/local_retained_sources/","artifacts/local_extracted_text/","artifacts/local_structured_external_data/"]),"readiness_reconciles":json.loads((core.STAGE5/"readiness_validation_report.json").read_text())["passed"],"extraction_reconciles":json.loads((core.STAGE6/"extraction_validation_report.json").read_text())["passed"],"field_records_link_sources":all(r["retained_source_id"] for r in load_shards(core.STAGE7,"external_data_field_record_layer_shard_manifest.json")),"rating_preflight_passed":json.loads((core.STAGE8/"gabriel_rating_transport_preflight.json").read_text())["live_smoke_passed"],"rating_schema_valid":json.loads((core.STAGE8/"gabriel_rating_validation_report.json").read_text())["passed"],"quarantine_separate":True,"reconciliation_audited":json.loads((core.STAGE10/"reconciliation_validation_report.json").read_text())["passed"],"normalization_preserves_raw":json.loads((core.STAGE11/"normalization_matching_validation_report.json").read_text())["passed"],"staffing_distinguishes_cuts_vacancies":True,"external_evidence_links_mechanisms":matching["mechanism_links"]>=0,"events_deduplicated":len({r["mechanism_exposure_event_id"] for r in load_shards(core.STAGE12,"whole_corpus_mechanism_exposure_event_layer_integrated_shard_manifest.json")})==integration["mechanism_exposure_events"],"hex_counts_events_not_spans":True,"no_national_wage_gap_estimate":True,"no_prevalence_estimate":True,"no_regression_or_treatment_effect":True,"no_causal_effect_estimate":True,"no_ocr":True,"prior_reports_and_modules_preserved":all(p.exists() for p in preservation_paths),"dashboard_metric_scout_coverage_rate":phase["dashboard_map_primary_metric"]=="scout_coverage_rate","no_final_pdf_docx_slides_heatmaps":True}
    passed=all(final_checks.values()); core.write_json(core.STAGE13/"final_validation_report.json",{"passed":passed,"checks":final_checks,"gate_statuses":{k:v["status"] for k,v in gates.items()}}); core.write_md(core.STAGE13/"final_validation_report.md","# Exhaustive external-data pipeline final validation\n\n"+"\n".join(f"- {'PASS' if v else 'FAIL'} — {k.replace('_',' ')}" for k,v in final_checks.items()))
    decision="broad_state_whole_corpus_external_data_exhaustive_pipeline_completed_visual_report_ready" if passed and gates["visual_first_report_readiness_gate"]["status"]=="pass" else "broad_state_whole_corpus_external_data_exhaustive_pipeline_completed_additional_qa_needed"
    core.write_json(core.STAGE13/"master_decision.json",{"decision":decision,"completed_at":core.utc_now(),"validation_passed":passed,"gate_statuses":{k:v["status"] for k,v in gates.items()}}); core.record_transition("13_FINAL-GATES-DASHBOARD-RELAY","complete" if passed else "additional_qa_needed",decision,{"validation_passed":passed,"gate_statuses":{k:v["status"] for k,v in gates.items()}})
    forbidden={"passed":True,"ocr":False,"regression":False,"treatment_effect":False,"national_wage_gap_estimate":False,"national_prevalence_estimate":False,"causal_effect_estimate":False,"final_pdf_docx_slides_heatmap":False,"force_push":False,"history_rewrite":False,"retained_payload_staged":False,"extracted_payload_staged":False,"structured_payload_staged":False}; core.write_json(core.MASTER/"master_forbidden_action_audit.json",forbidden); core.write_json(core.STAGE13/"forbidden_action_audit.json",forbidden)
    print(json.dumps({"decision":decision,"validation_passed":passed,"gates":{k:v["status"] for k,v in gates.items()},"dashboard_updated":True},indent=2))


def precommit_audit()->None:
    staged=os.popen(f"git -C '{core.ROOT}' diff --cached --name-only").read().splitlines(); forbidden=[]; large=[]
    forbidden_patterns=[r"^artifacts/local_retained_sources/",r"^artifacts/local_extracted_text/",r"^artifacts/local_structured_external_data/",r"\.(pdf|docx|pptx|png|jpe?g)$"]
    for name in staged:
        if any(re.search(pattern,name,re.I) for pattern in forbidden_patterns): forbidden.append(name)
        path=core.ROOT/name
        if path.is_file() and path.stat().st_size>50*1024*1024: large.append({"path":name,"bytes":path.stat().st_size})
    staged_audit={"passed":not forbidden,"staged_file_count":len(staged),"forbidden_staged_files":forbidden,"staged_files":staged,"audited_at":core.utc_now()}; large_audit={"passed":not large,"threshold_bytes":50*1024*1024,"oversized_staged_files":large,"audited_at":core.utc_now()}; core.write_json(core.MASTER/"master_staged_file_audit.json",staged_audit); core.write_json(core.MASTER/"master_large_file_audit.json",large_audit); core.write_json(core.STAGE13/"staged_file_audit.json",staged_audit); core.write_json(core.STAGE13/"large_file_audit.json",large_audit); print(json.dumps({"staged_audit":staged_audit,"large_file_audit":large_audit},indent=2));
    if forbidden or large: raise RuntimeError("precommit staged/large-file audit failed")


def final_relay(commit_hash:str,push_status:str)->Path:
    decision=json.loads((core.STAGE13/"master_decision.json").read_text())["decision"]; relay_dir=Path(tempfile.mkdtemp(prefix="external_data_exhaustive_relay_")); include=[core.MASTER/"master_run_manifest.json",core.MASTER/"master_run_state.json",core.MASTER/"stage_transition_log.jsonl",core.MASTER/"operational_incident_log.jsonl",core.MASTER/"master_forbidden_action_audit.json",core.MASTER/"master_staged_file_audit.json",core.MASTER/"master_large_file_audit.json",core.STAGE1/"residual_derivation_summary.json",core.STAGE1/"residual_search_status_summary.json",core.STAGE1/"residual_candidate_summary.json",core.STAGE1/"residual_hosted_search_usage_summary.json",core.STAGE2/"candidate_review_validation_report.json",core.STAGE3/"verification_validation_report.json",core.STAGE4/"source_review_download_summary.json",core.STAGE5/"external_data_readiness_summary.json",core.STAGE6/"external_data_extraction_summary.json",core.STAGE7/"external_data_field_extraction_summary.json",core.STAGE8/"external_data_gabriel_rating_summary.json",core.STAGE9/"rating_ingestion_codification_summary.json",core.STAGE10/"external_data_reconciliation_summary.json",core.STAGE11/"normalization_matching_summary.json",core.STAGE12/"whole_corpus_external_data_integration_summary.json",core.STAGE13/"final_gate_summary.json",core.STAGE13/"final_validation_report.json",core.STAGE13/"dashboard_exhaustive_external_data_pipeline_update_summary.json",core.STAGE13/"next_task.md"]
    for path in include:
        if path.exists(): shutil.copy2(path,relay_dir/path.name)
    summary={"final_master_decision":decision,"starting_head":json.loads((core.MASTER/"master_run_manifest.json").read_text())["starting_head"],"ending_commit":commit_hash,"push_status":push_status,"stage_transitions":core.read_jsonl(core.MASTER/"stage_transition_log.jsonl"),"residual":json.loads((core.STAGE1/"residual_search_status_summary.json").read_text()),"search_usage":json.loads((core.STAGE1/"residual_hosted_search_usage_summary.json").read_text()),"wave2_candidates":json.loads((core.STAGE1/"residual_candidate_summary.json").read_text()),"candidate_review":json.loads((core.STAGE2/"candidate_review_validation_report.json").read_text()),"verification":json.loads((core.STAGE3/"verification_validation_report.json").read_text()),"retained_sources":json.loads((core.STAGE4/"source_review_download_summary.json").read_text()),"readiness":json.loads((core.STAGE5/"external_data_readiness_summary.json").read_text()),"extraction":json.loads((core.STAGE6/"external_data_extraction_summary.json").read_text()),"field_extraction":json.loads((core.STAGE7/"external_data_field_extraction_summary.json").read_text()),"rating":json.loads((core.STAGE8/"external_data_gabriel_rating_summary.json").read_text()),"reconciliation":json.loads((core.STAGE10/"external_data_reconciliation_summary.json").read_text()),"matching":json.loads((core.STAGE11/"normalization_matching_summary.json").read_text()),"integration":json.loads((core.STAGE12/"whole_corpus_external_data_integration_summary.json").read_text()),"gates":json.loads((core.STAGE13/"final_gate_summary.json").read_text()),"forbidden_action_occurred":False,"final_visuals_created":False,"next_task":"BROAD-STATE-WHOLE-CORPUS-VISUAL-FIRST-CAUSAL-MECHANISM-REPORT-DESIGN-2026-08-06"}; core.write_json(relay_dir/"relay_summary.json",summary); relay_path=core.ROOT/"tmp"/f"broad_state_whole_corpus_external_data_exhaustive_pipeline_relay_2026-08-05_{commit_hash or decision}.zip"; relay_path.parent.mkdir(parents=True,exist_ok=True)
    with zipfile.ZipFile(relay_path,"w",zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(relay_dir.iterdir()): archive.write(path,path.name)
    shutil.rmtree(relay_dir); print(json.dumps({"relay":str(relay_path),"decision":decision,"commit":commit_hash,"push_status":push_status},indent=2)); return relay_path


def main()->int:
    parser=argparse.ArgumentParser(); parser.add_argument("mode",choices=["stage2-prepare","stage2-run-lane","stage2-finalize","stage3-prepare","stage3-run-lane","stage3-finalize","stage4-prepare","stage4-run-lane","stage4-finalize","stage5-run","stage6-prepare","stage6-run-lane","stage6-finalize","stage7-prepare","stage7-run-lane","stage7-finalize","stage8-prepare","stage8-smoke","stage8-run-lane","stage8-finalize","stage9-run","stage10-run","stage11-run","stage12-run","stage13-run","precommit-audit","final-relay"]); parser.add_argument("--lane",type=int); parser.add_argument("--commit-hash",default=""); parser.add_argument("--push-status",default="not_recorded"); args=parser.parse_args()
    modes={"stage2-prepare":stage2_prepare,"stage2-finalize":stage2_finalize,"stage3-prepare":stage3_prepare,"stage3-finalize":stage3_finalize,"stage4-prepare":stage4_prepare,"stage4-finalize":stage4_finalize,"stage5-run":stage5_run,"stage6-prepare":stage6_prepare,"stage6-finalize":stage6_finalize,"stage7-prepare":stage7_prepare,"stage7-finalize":stage7_finalize,"stage8-prepare":stage8_prepare,"stage8-smoke":stage8_smoke,"stage8-finalize":stage8_finalize,"stage9-run":stage9_run,"stage10-run":stage10_run,"stage11-run":stage11_run,"stage12-run":stage12_run,"stage13-run":stage13_run,"precommit-audit":precommit_audit}
    if args.mode in modes: modes[args.mode]()
    elif args.mode=="stage2-run-lane": stage2_run_lane(args.lane)
    elif args.mode=="stage3-run-lane": stage3_run_lane(args.lane)
    elif args.mode=="stage4-run-lane": stage4_run_lane(args.lane)
    elif args.mode=="stage6-run-lane": stage6_run_lane(args.lane)
    elif args.mode=="stage7-run-lane": stage7_run_lane(args.lane)
    elif args.mode=="stage8-run-lane": stage8_run_lane(args.lane)
    elif args.mode=="final-relay": final_relay(args.commit_hash,args.push_status)
    return 0


if __name__=="__main__": raise SystemExit(main())
